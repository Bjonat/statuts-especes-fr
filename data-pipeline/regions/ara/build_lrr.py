#!/usr/bin/env python3
"""Listes rouges régionales unifiées Auvergne-Rhône-Alpes 2024 — vertébrés."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

PRODUCER = (
    "DREAL Auvergne-Rhône-Alpes / CSRPN Auvergne-Rhône-Alpes / "
    "LPO Auvergne-Rhône-Alpes / Observatoire régional de la biodiversité / partenaires"
)
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[a-z]{0,3})$")
NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}
TABLE_NS = NS["table"]

FILES = {
    "amph_rept_chiro": {
        "filename": "vertebres-amph-rept-chiro.xlsx",
        "sha256": "ae49929b0a3d226fa392850c4fa95d928d5f374f179a01fd9ea5f71feac1a581",
        "sourceUrl": (
            "https://www.biodiversite-auvergne-rhone-alpes.fr/wp-content/uploads/2023/03/"
            "LR_AURA2024_Chauves-souris_reptiles_amphibiens.xlsx"
        ),
        "landingPage": (
            "https://www.auvergne-rhone-alpes.developpement-durable.gouv.fr/"
            "2024-08-liste-rouge-des-amphibiens-reptiles-et-a26033.html"
        ),
    },
    "oiseaux_mamm": {
        "filename": "oiseaux-mammiferes.ods",
        "sha256": "3308ae670319c729f248d444ddfb08b621a02cbc52610c3e4ad2a548eefacd7b",
        "sourceUrl": (
            "https://www.auvergne-rhone-alpes.developpement-durable.gouv.fr/"
            "IMG/ods/2024-lrr-oisx_mamm_web-dreal.ods"
        ),
        "landingPage": (
            "https://www.auvergne-rhone-alpes.developpement-durable.gouv.fr/"
            "2024-05-liste-rouge-oiseaux-nicheurs-et-mammiferes-a25597.html"
        ),
    },
}

SOURCES = [
    {
        "key": "amphibiens",
        "file_key": "amph_rept_chiro",
        "group": "Amphibiens",
        "id": "dreal-ara-lrr-amphibiens-2024",
        "name": "Liste rouge Amphibiens Auvergne-Rhône-Alpes",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "reptiles",
        "file_key": "amph_rept_chiro",
        "group": "Reptiles",
        "id": "dreal-ara-lrr-reptiles-2024",
        "name": "Liste rouge Reptiles Auvergne-Rhône-Alpes",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "chiropteres",
        "file_key": "amph_rept_chiro",
        "group": "Chauves-souris",
        "id": "dreal-ara-lrr-chiropteres-2024",
        "name": "Liste rouge Chiroptères Auvergne-Rhône-Alpes",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "oiseaux-nicheurs",
        "file_key": "oiseaux_mamm",
        "sheet": "LRR-oiseaux-nich",
        "id": "dreal-ara-lrr-oiseaux-nicheurs-2024",
        "name": "Liste rouge Oiseaux nicheurs Auvergne-Rhône-Alpes",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "mammiferes",
        "file_key": "oiseaux_mamm",
        "sheet": "LRR_mamm-terr",
        "id": "dreal-ara-lrr-mammiferes-2024",
        "name": "Liste rouge Mammifères terrestres (hors chiroptères) Auvergne-Rhône-Alpes",
        "version": "2024",
        "year": 2024,
    },
]


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip().casefold()


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def normalize_category(value: object) -> str | None:
    text = clean(value)
    match = re.fullmatch(r"(?i)(EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA)([A-Za-z]{0,3})", text)
    if not match:
        return None
    base = match.group(1).upper()
    if base.startswith("CR") and "*" in match.group(1):
        base = "CR*"
    suffix = match.group(2)
    if base == "NA" and suffix:
        category = "NA" + suffix.lower()
    elif suffix:
        return None
    else:
        category = base
    if VALID_LRR_CATEGORY.fullmatch(category):
        return category
    return None


def read_amph_rept_chiro(path: Path) -> dict[str, list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"{path.name}: feuille vide")
    header = [clean(value) for value in rows[0]]
    expected = ["Groupe", "cd_nom", "Nom vernaculaire", "Nom scientifique", "LR AuRA 2024"]
    if header[:5] != expected:
        raise RuntimeError(f"{path.name}: en-têtes inattendus {header[:5]}")
    by_group: dict[str, list[dict]] = defaultdict(list)
    for values in rows[1:]:
        if not values or not values[0]:
            continue
        group = clean(values[0])
        category = normalize_category(values[4])
        if category is None:
            continue
        by_group[group].append(
            {
                "code": as_int(values[1]),
                "name": clean(values[3]),
                "category": category,
            }
        )
    return by_group


def ods_sheet_rows(path: Path, sheet_name: str) -> list[list[str]]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))
    for table in root.findall(".//table:table", NS):
        if table.get(f"{{{TABLE_NS}}}name") != sheet_name:
            continue
        rows: list[list[str]] = []
        for table_row in table.findall("table:table-row", NS):
            cells: list[str] = []
            for cell in table_row.findall("table:table-cell", NS):
                repeat = int(cell.get(f"{{{TABLE_NS}}}number-columns-repeated") or 1)
                text = clean("".join(cell.itertext()))
                cells.extend([text] * min(repeat, 20))
            if any(cells):
                rows.append(cells)
        return rows
    raise RuntimeError(f"{path.name}: onglet introuvable {sheet_name}")


def read_oiseaux(path: Path) -> list[dict]:
    rows = ods_sheet_rows(path, "LRR-oiseaux-nich")
    header = rows[0]
    if header[:4] != ["cd_nom", "Nom français", "Nom latin", "Statuts"]:
        raise RuntimeError(f"{path.name}: en-têtes oiseaux inattendus {header[:4]}")
    parsed = []
    for values in rows[1:]:
        category = normalize_category(values[3] if len(values) > 3 else "")
        if category is None:
            continue
        parsed.append(
            {
                "code": as_int(values[0]),
                "name": clean(values[2] if len(values) > 2 else ""),
                "category": category,
            }
        )
    return parsed


def read_mammiferes(path: Path) -> list[dict]:
    rows = ods_sheet_rows(path, "LRR_mamm-terr")
    header = rows[0]
    if header[:3] != ["Nom français", "Nom latin", "Statut"]:
        raise RuntimeError(f"{path.name}: en-têtes mammifères inattendus {header[:3]}")
    parsed = []
    for values in rows[1:]:
        category = normalize_category(values[2] if len(values) > 2 else "")
        if category is None:
            continue
        parsed.append(
            {
                "code": None,
                "name": clean(values[1] if len(values) > 1 else ""),
                "category": category,
            }
        )
    return parsed


def parse_all(input_dir: Path) -> dict[str, list[dict]]:
    for meta in FILES.values():
        path = input_dir / meta["filename"]
        actual = sha256(path)
        if actual != meta["sha256"]:
            raise SystemExit(f"{path.name}: SHA-256 inattendu {actual} != {meta['sha256']}")

    amph_groups = read_amph_rept_chiro(input_dir / FILES["amph_rept_chiro"]["filename"])
    oiseaux_path = input_dir / FILES["oiseaux_mamm"]["filename"]
    parsed: dict[str, list[dict]] = {}
    for source in SOURCES:
        if source["key"] in {"amphibiens", "reptiles", "chiropteres"}:
            rows = amph_groups.get(source["group"], [])
            if not rows:
                raise RuntimeError(f"Groupe absent: {source['group']}")
            parsed[source["key"]] = rows
        elif source["key"] == "oiseaux-nicheurs":
            parsed[source["key"]] = read_oiseaux(oiseaux_path)
        else:
            parsed[source["key"]] = read_mammiferes(oiseaux_path)
    return parsed


def wanted_from_parsed(parsed: dict[str, list[dict]]):
    codes: set[int] = set()
    names: set[str] = set()
    for rows in parsed.values():
        for row in rows:
            if row["code"] is not None:
                codes.add(row["code"])
            if row["name"]:
                names.add(normalize(row["name"]))
    return codes, names


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str]):
    by_cd_nom: dict[int, tuple[int, str | None]] = {}
    by_name = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom_raw = str(row.get("CD_NOM") or "").strip()
            cd_ref_raw = str(row.get("CD_REF") or "").strip()
            if not cd_nom_raw.isdigit() or not cd_ref_raw.isdigit():
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            cd_nom = int(cd_nom_raw)
            cd_ref = int(cd_ref_raw)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = (cd_ref, realm)
            if realm and wanted_names:
                for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                    label = row.get(field)
                    if label and normalize(label) in wanted_names:
                        by_name[normalize(label)].add((cd_ref, realm))
    return by_cd_nom, by_name


def resolve(row, by_cd_nom, by_name):
    code = row["code"]
    if code is not None and code in by_cd_nom:
        cd_ref, realm = by_cd_nom[code]
        if realm:
            return cd_ref, realm, "cd_nom"
        return None, None, "excluded_realm"
    if row["name"]:
        candidates = by_name.get(normalize(row["name"]), set())
        if len(candidates) == 1:
            cd_ref, realm = next(iter(candidates))
            return cd_ref, realm, "name"
        if len(candidates) > 1:
            return None, None, "ambiguous"
    return None, None, "unmatched"


def build_package(source, rows, input_dir: Path, by_cd_nom, by_name, checked_at: str):
    stats = {
        "rows": len(rows),
        "matched": 0,
        "cd_nom": 0,
        "name": 0,
        "excluded_realm": 0,
        "unmatched": 0,
        "ambiguous": 0,
        "unexpectedRealm": 0,
        "unresolvedSample": [],
        "values": {},
    }
    values = defaultdict(int)
    statuses = []
    seen = set()

    for row in rows:
        cd_ref, realm, mode = resolve(row, by_cd_nom, by_name)
        if cd_ref is None or realm is None:
            stats[mode] += 1
            if len(stats["unresolvedSample"]) < 30:
                stats["unresolvedSample"].append(
                    {
                        "code": row["code"],
                        "taxon": row["name"],
                        "category": row["category"],
                        "reason": mode,
                    }
                )
            continue
        if realm != "fauna":
            stats["unexpectedRealm"] += 1
            continue
        stats["matched"] += 1
        stats[mode] += 1
        values[row["category"]] += 1
        key = (cd_ref, row["category"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": cd_ref,
                "region": "ARA",
                "category": "red_list_regional",
                "label": "Liste rouge régionale",
                "value": row["category"],
                "sourceId": source["id"],
                "scope": "regional",
            }
        )

    candidates = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    file_meta = FILES[source["file_key"]]
    file_path = input_dir / file_meta["filename"]
    covered_refs = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": sha256(file_path),
            "landingPage": file_meta["landingPage"],
            "sourceUrl": file_meta["sourceUrl"],
        },
        "replaces": [
            {
                "region": "ARA",
                "category": "red_list_regional",
                "realm": "fauna",
                "cdRefs": covered_refs,
            },
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": stats,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    parsed = parse_all(input_dir)
    wanted_codes, wanted_names = wanted_from_parsed(parsed)
    by_cd_nom, by_name = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names)

    total_statuses = 0
    for source in SOURCES:
        package = build_package(
            source, parsed[source["key"]], input_dir, by_cd_nom, by_name, args.checked_at
        )
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"ara-lrr-{source['key']}.json"
        output.write_text(
            json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n",
            encoding="utf-8",
        )
        total_statuses += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")

    print(f"Auvergne-Rhône-Alpes LRR: {len(SOURCES)} paquets, {total_statuses} statuts")


if __name__ == "__main__":
    main()
