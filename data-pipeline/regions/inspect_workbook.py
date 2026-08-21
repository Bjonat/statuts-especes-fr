#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import sys
from pathlib import Path


def non_empty_rows(rows):
    return [list(row) for row in rows if any(str(value or '').strip() for value in row)]


def inspect_xlsx(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    print(f"Onglets ({len(workbook.sheetnames)}): {' | '.join(workbook.sheetnames)}")
    for name in workbook.sheetnames:
        sheet = workbook[name]
        rows = non_empty_rows(sheet.iter_rows(values_only=True))
        print(f"\n### {name} - {len(rows)} lignes non vides")
        for row in rows[:15]:
            print(json.dumps(row, ensure_ascii=False, default=str))


def inspect_xls(path: Path) -> None:
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True)
    print(f"Onglets ({workbook.nsheets}): {' | '.join(workbook.sheet_names())}")
    for name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(name)
        rows = non_empty_rows(sheet.row_values(index) for index in range(sheet.nrows))
        print(f"\n### {name} - {len(rows)} lignes non vides")
        for row in rows[:15]:
            print(json.dumps(row, ensure_ascii=False, default=str))


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit('Usage: inspect_workbook.py <fichier.xls|xlsx>')

    path = Path(sys.argv[1])
    payload = path.read_bytes()
    print(f"Fichier: {path.name} - {len(payload):,} octets")
    print(f"Signature hex: {payload[:16].hex()}")

    preview = payload[:256].lower()
    if b'<html' in preview or b'<!doctype html' in preview:
        print(payload[:800].decode('utf-8', errors='replace').replace('\n', ' '))
        raise SystemExit('Le serveur a renvoyé du HTML au lieu d’un tableur')

    suffix = path.suffix.lower()
    if suffix == '.xlsx' or payload.startswith(b'PK'):
        inspect_xlsx(path)
    elif suffix == '.xls' or payload.startswith(bytes.fromhex('d0cf11e0a1b11ae1')):
        inspect_xls(path)
    else:
        raise SystemExit(f'Format de classeur non reconnu: {suffix}')


if __name__ == '__main__':
    main()
