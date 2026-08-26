#!/usr/bin/env python3
"""ZNIEFF Occitanie — flore (vasculaire, bryophytes, characées) + faune."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LANDING_URL = (
    "https://www.occitanie.developpement-durable.gouv.fr/"
    "vers-des-znieff-troisieme-generation-en-occitanie-a24635.html"
)
PRODUCER = "DREAL Occitanie / CSRPN Occitanie / partenaires naturalistes"
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
MAX_VALUE_LENGTH = 80

FLORA_ZONES = [
    {
        "keys": ("det_med_occ", "det_méditerranée"),
        "header_prefixes": ("Det_Med_Occ", "Det_med_Occ"),
        "label": "Méditerranée",
    },
    {
        "keys": ("det_mc_occ",),
        "header_prefixes": ("Det_MC_Occ",),
        "label": "Massif central",
    },
    {
        "keys": ("det_pyr_occ",),
        "header_prefixes": ("Det_Pyr_Occ",),
        "label": "Pyrénées",
    },
    {
        "keys": ("det_ba_occ",),
        "header_prefixes": ("Det_BA_Occ",),
        "label": "Bassin aquitain",
    },
]

FAUNA_ZONE_NAMES = {
    "so": "Sud-Ouest",
    "med": "Méditerranée",
    "pyr": "Pyrénées",
    "mc": "Massif central",
    "oc": "Occitanie",
}

FILES = {
    "flora": {
        "filename": "znieff-flora.xlsx",
        "sha256": "87464cbb51ccc07de54586d10c6071b0a5344027f8c335dea2f06fcb877bb834",
        "sourceUrl": (
            "https://www.occitanie.developpement-durable.gouv.fr/IMG/xlsx/"
            "liste_taxons_det_flore_occitanie_cotation_v13-v16_osmose_public.xlsx"
        ),
    },
    "fauna": {
        "filename": "znieff-fauna.xlsx",
        "sha256": "ec66eed10fde0e97558c1f2a973fd8480037b722ec4e03814517e5944754d873",
        "sourceUrl": (
            "https://www.occitanie.developpement-durable.gouv.fr/IMG/xlsx/"
            "listes_faune_znieff_20240725.xlsx"
        ),
    },
}

SOURCES = [
    {
        "key": "flora",
        "file_key": "flora",
        "sheet": "Flore Dét",
        "id": "dreal-occ-znieff-flora-2023",
        "name": "ZNIEFF Flore vasculaire Occitanie",
        "version": "2023-02 CSRPN / taxref v16",
        "year": 2023,
        "realm": "flora",
    },
    {
        "key": "bryophytes",
        "file_key": "flora",
        "sheet": "Bryo Dét",
        "id": "dreal-occ-znieff-bryophytes-2023",
        "name": "ZNIEFF Bryophytes Occitanie",
        "version": "2023-02 CSRPN / taxref v16",
        "year": 2023,
        "realm": "flora",
    },
    {
        "key": "characees",
        "file_key": "flora",
        "sheet": "Chara Dét",
        "id": "dreal-occ-znieff-characees-2023",
        "name": "ZNIEFF Characées Occitanie",
        "version": "2023-02 CSRPN / taxref v16",
        "year": 2023,
        "realm": "flora",
    },
    {
        "key": "fauna",
        "file_key": "fauna",
        "sheet": "Faune Occitanie",
        "id": "dreal-occ-znieff-fauna-2024-07",
        "name": "ZNIEFF Faune Occitanie",
        "version": "2024-07-25",
        "year": 2024,
        "realm": "fauna",
    },
]


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").replace("–", "-").replace("—", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def find_det_columns(headers: list[str]) -> list[tuple[int, str]]:
    found: list[tuple[int, str]] = []
    normalized_headers = [normalize(header) for header in headers]
    for zone in FLORA_ZONES:
        index = None
        for candidate in zone["header_prefixes"]:
            target = normalize(candidate)
            for header_index, header_key in enumerate(normalized_headers):
                if header_key.startswith(target):
                    index = header_index
                    break
            if index is not None:
                break
        if index is None:
            raise RuntimeError(f"Colonne de déterminance introuvable pour {zone['label']}: {headers}")
        found.append((index, zone["label"]))
    return found


def read_flora_sheet(path: Path, sheet_name: str) -> tuple[list[dict], set[int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(f"Onglet absent: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if len(rows) < 3:
        raise RuntimeError(f"{sheet_name}: trop peu de lignes")
    headers = [clean(value) for value in rows[1]]
    code_index = next(
        (index for index, header in enumerate(headers) if normalize(header).startswith("cd_nom")),
        None,
    )
    name_index = next(
        (
            index
            for index, header in enumerate(headers)
            if "nom retenu" in normalize(header) or "nom scientifique" in normalize(header)
        ),
        None,
    )
    if code_index is None:
        raise RuntimeError(f"{sheet_name}: colonne cd_nom absente")
    det_columns = find_det_columns(headers)
    statuses: list[dict] = []
    evaluated: set[int] = set()
    for values in rows[2:]:
        code = as_int(values[code_index] if code_index < len(values) else None)
        if code is None:
            continue
        name = clean(values[name_index]) if name_index is not None and name_index < len(values) else ""
        saw_evaluation = False
        for det_index, zone_label in det_columns:
            raw = clean(values[det_index] if det_index < len(values) else "")
            if not raw:
                continue
            key = normalize(raw)
            if key not in {"d", "nd"}:
                continue
            saw_evaluation = True
            if key == "d":
                statuses.append(
                    {
                        "code": code,
                        "name": name,
                        "value": "Déterminante",
                        "scope": "partial",
                        "scopeLabel": zone_label,
                    }
                )
        if saw_evaluation:
            evaluated.add(code)
    return statuses, evaluated


def normalize_fauna_scope(raw: object) -> tuple[str, str | None]:
    text = clean(raw)
    if not text:
        return "regional", None
    key = normalize(text)
    if key == "occitanie":
        return "regional", None
    # Extract zone tokens SO / MED / PYR / MC / OC
    tokens = re.findall(r"\b(so|med|pyr|mc|oc)\b", key)
    labels = []
    seen = set()
    for token in tokens:
        label = FAUNA_ZONE_NAMES[token]
        if label not in seen:
            seen.add(label)
            labels.append(label)
    if not labels:
        # Fallback: keep cleaned source text if compact enough
        if len(text) <= MAX_VALUE_LENGTH:
            return "partial", text
        return "regional", None
    if labels == ["Occitanie"]:
        return "regional", None
    return "partial", " / ".join(labels)


def read_fauna_sheet(path: Path) -> tuple[list[dict], set[int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Faune Occitanie"]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if len(rows) < 3:
        raise RuntimeError("Faune Occitanie: trop peu de lignes")
    headers = [clean(value) for value in rows[1]]
    if headers[:4] != ["Groupe", "Nom scientifique", "Nom vernaculaire", "CD_NOM"]:
        raise RuntimeError(f"En-têtes faune inattendus: {headers[:4]}")
    statuses: list[dict] = []
    evaluated: set[int] = set()
    for values in rows[2:]:
        code = as_int(values[3] if len(values) > 3 else None)
        if code is None:
            continue
        name = clean(values[1] if len(values) > 1 else "")
        scope, scope_label = normalize_fauna_scope(values[9] if len(values) > 9 else "")
        evaluated.add(code)
        entry = {
            "code": code,
            "name": name,
            "value": "Déterminante",
            "scope": scope,
        }
        if scope_label:
            entry["scopeLabel"] = scope_label
        statuses.append(entry)
    return statuses, evaluated


def parse_all(input_dir: Path):
    for meta in FILES.values():
        path = input_dir / meta["filename"]
        actual = sha256(path)
        if actual != meta["sha256"]:
            raise SystemExit(f"{path.name}: SHA-256 inattendu {actual} != {meta['sha256']}")

    parsed: dict[str, dict] = {}
    for source in SOURCES:
        file_path = input_dir / FILES[source["file_key"]]["filename"]
        if source["key"] == "fauna":
            rows, evaluated = read_fauna_sheet(file_path)
        else:
            rows, evaluated = read_flora_sheet(file_path, source["sheet"])
        if not rows:
            raise SystemExit(f"{source['id']}: aucun statut positif")
        parsed[source["key"]] = {"rows": rows, "evaluated": evaluated}
    return parsed


def taxref_lookup(path: Path, wanted_codes: set[int]):
    by_cd_nom: dict[int, tuple[int, str | None]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom_raw = clean(row.get("CD_NOM"))
            cd_ref_raw = clean(row.get("CD_REF"))
            if not cd_nom_raw.isdigit() or not cd_ref_raw.isdigit():
                continue
            cd_nom = int(cd_nom_raw)
            if cd_nom not in wanted_codes:
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            by_cd_nom[cd_nom] = (int(cd_ref_raw), realm)
    return by_cd_nom


def build_package(source, payload, input_dir: Path, by_cd_nom, checked_at: str):
    stats = {
        "rows": len(payload["rows"]),
        "matched": 0,
        "unmatched": 0,
        "unexpectedRealm": 0,
        "excluded_realm": 0,
        "unresolvedSample": [],
        "values": {},
        "scopes": {},
    }
    values = defaultdict(int)
    scopes = defaultdict(int)
    statuses = []
    seen = set()
    matched_refs: set[int] = set()
    evaluated_refs: set[int] = set()

    for code in payload["evaluated"]:
        resolved = by_cd_nom.get(code)
        if resolved and resolved[1] == source["realm"]:
            evaluated_refs.add(resolved[0])

    for row in payload["rows"]:
        resolved = by_cd_nom.get(row["code"])
        if resolved is None:
            stats["unmatched"] += 1
            if len(stats["unresolvedSample"]) < 30:
                stats["unresolvedSample"].append(
                    {"code": row["code"], "taxon": row["name"], "reason": "unmatched"}
                )
            continue
        cd_ref, realm = resolved
        if realm is None:
            stats["excluded_realm"] += 1
            continue
        if realm != source["realm"]:
            stats["unexpectedRealm"] += 1
            continue
        stats["matched"] += 1
        values[row["value"]] += 1
        scope_key = row.get("scopeLabel") or row["scope"]
        scopes[scope_key] += 1
        key = (cd_ref, row["value"], row["scope"], row.get("scopeLabel"))
        if key in seen:
            continue
        seen.add(key)
        matched_refs.add(cd_ref)
        status = {
            "cdRef": cd_ref,
            "region": "OCC",
            "category": "znieff",
            "label": "Déterminante ZNIEFF",
            "value": row["value"],
            "sourceId": source["id"],
            "scope": row["scope"],
        }
        if row.get("scopeLabel"):
            status["scopeLabel"] = row["scopeLabel"]
        statuses.append(status)

    candidates = stats["matched"] + stats["unmatched"]
    # Match rate against positive rows (each zone row counts)
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    stats["scopes"] = dict(sorted(scopes.items()))
    file_meta = FILES[source["file_key"]]
    covered = sorted(evaluated_refs | matched_refs)
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": sha256(input_dir / file_meta["filename"]),
            "landingPage": LANDING_URL,
            "sourceUrl": file_meta["sourceUrl"],
        },
        "replaces": [
            {
                "region": "OCC",
                "category": "znieff",
                "realm": source["realm"],
                "cdRefs": covered,
            },
        ],
        "statuses": sorted(
            statuses,
            key=lambda status: (
                status["cdRef"],
                status.get("scopeLabel") or "",
                status["value"],
            ),
        ),
        "diagnostics": stats,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    parsed = parse_all(input_dir)
    wanted_codes: set[int] = set()
    for payload in parsed.values():
        wanted_codes.update(payload["evaluated"])
        for row in payload["rows"]:
            wanted_codes.add(row["code"])
    by_cd_nom = taxref_lookup(Path(args.taxref), wanted_codes)

    total = 0
    for source in SOURCES:
        package = build_package(
            source, parsed[source["key"]], input_dir, by_cd_nom, args.checked_at
        )
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"occ-znieff-{source['key']}.json"
        output.write_text(
            json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n",
            encoding="utf-8",
        )
        total += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")

    print(f"Occitanie ZNIEFF: {len(SOURCES)} paquets, {total} statuts")


if __name__ == "__main__":
    main()
