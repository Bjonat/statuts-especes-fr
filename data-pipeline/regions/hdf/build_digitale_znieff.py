#!/usr/bin/env python3
"""ZNIEFF Hauts-de-France : flore/bryophytes Digitale + faune historique DREAL."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LANDING_URL = "https://www.cbnhdf.fr/je-telecharge"
PRODUCER = "Conservatoire botanique national des Hauts-de-France / CSRPN Hauts-de-France"
SCOPE_LABEL = "Hauts-de-France"
TERRITORY = "HDF"
# [Oui] = erreur / douteux / cultivé uniquement — hors publication déterminante.
VALUE_BY_CODE = {
    "Oui": "Oui",
    "(Oui)": "Oui (disparu/présumé)",
    "pp": "Oui (pro parte)",
}
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
MAX_VALUE_LENGTH = 80

SOURCES = [
    {
        "key": "flora",
        "filename": "digitale-flora.xlsx",
        "sheet": "REG-DIGITALE-BS-BIF-FVF-PV_4.0",
        "id": "cbnhdf-digitale-znieff-hdf-flora-2026-03-31",
        "name": "Espèces déterminantes ZNIEFF flore vasculaire Hauts-de-France",
        "version": "Digitale BS-BIF-FVF-PV 4.0 (2026-03-31)",
        "sourceUrl": "https://www.cbnhdf.fr/system/files/2026-05/DIGITALE_BS-BIF-FVF_PV_4.0_20260331.xlsx",
        "sha256": "71ae71b770f7b3911349e501caaaa65ac7dba8172d12b96ef4b90d5056995c95",
        "minVolume": 800,
        "outName": "hdf-znieff-flora.json",
    },
    {
        "key": "bryophytes",
        "filename": "digitale-bryophytes.xlsx",
        "sheet": "REG-DIGITALE-BS-BIF-FVF-MH_4.0",
        "id": "cbnhdf-digitale-znieff-hdf-bryophytes-2026-03-31",
        "name": "Espèces déterminantes ZNIEFF bryophytes Hauts-de-France",
        "version": "Digitale BS-BIF-FVF-MH 4.0 (2026-03-31)",
        "sourceUrl": "https://www.cbnhdf.fr/system/files/2026-05/DIGITALE_BS-BIF-FVF_MH_4.0_20260331.xlsx",
        "sha256": "810cc4cc9458721710a826d009884698fcf9b06d059af41153197c12470cb3bc",
        "minVolume": 300,
        "outName": "hdf-znieff-bryophytes.json",
    },
]


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def read_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"{path.name}: feuille absente ({sheet_name})")
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        header = [clean(value) for value in next(iterator)]
        required = {"CH_Territoire", "CH_DetermZNIEFF", "CD_REF_TAXREF"}
        if not required.issubset(set(header)):
            raise SystemExit(f"{path.name}: colonnes manquantes {sorted(required - set(header))}")
        rows = []
        for values in iterator:
            row = {header[index]: values[index] if index < len(values) else None for index in range(len(header))}
            rows.append(row)
        return rows
    finally:
        workbook.close()


def select_rows(rows: list[dict[str, object]]):
    selected = []
    raw_codes: Counter = Counter()
    for row in rows:
        if clean(row.get("CH_Territoire")) != TERRITORY:
            continue
        code = clean(row.get("CH_DetermZNIEFF"))
        raw_codes[code or "<vide>"] += 1
        value = VALUE_BY_CODE.get(code)
        if not value:
            continue
        if len(value) > MAX_VALUE_LENGTH:
            continue
        cd_ref = as_int(row.get("CD_REF_TAXREF"))
        if cd_ref is None:
            continue
        selected.append(
            {
                "cd_ref": cd_ref,
                "value": value,
                "name": clean(row.get("CH_NomCompletTAXREF") or row.get("CH_NomComp")),
                "raw": code,
            }
        )
    return selected, raw_codes


def parse_all(input_dir: Path) -> dict[str, dict]:
    parsed = {}
    for source in SOURCES:
        path = input_dir / source["filename"]
        digest = sha256(path)
        if digest != source["sha256"]:
            raise SystemExit(f"{source['filename']}: SHA-256 Digitale inattendu: {digest}")
        rows = read_rows(path, source["sheet"])
        selected, raw_codes = select_rows(rows)
        parsed[source["key"]] = {
            "selected": selected,
            "raw_codes": raw_codes,
            "rows_source": len(rows),
            "digest": digest,
        }
    return parsed


def wanted_from_parsed(parsed: dict[str, dict]) -> set[int]:
    codes: set[int] = set()
    for entry in parsed.values():
        for row in entry["selected"]:
            codes.add(row["cd_ref"])
    return codes


def taxref_lookup(path: Path, wanted_codes: set[int]) -> dict[int, str | None]:
    by_cd_ref: dict[int, str | None] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_ref_raw = clean(row.get("CD_REF"))
            if not cd_ref_raw.isdigit():
                continue
            cd_ref = int(cd_ref_raw)
            if cd_ref not in wanted_codes:
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            by_cd_ref[cd_ref] = realm
    return by_cd_ref


def build_package(source: dict, entry: dict, by_cd_ref: dict[int, str | None], checked_at: str) -> dict:
    stats: Counter = Counter()
    stats["rows_source"] = entry["rows_source"]
    stats["rows_hdf_znieff"] = len(entry["selected"])
    values: Counter = Counter()
    statuses = []
    seen = set()
    unresolved = []

    for row in entry["selected"]:
        realm = by_cd_ref.get(row["cd_ref"])
        if realm is None:
            stats["unmatched"] += 1
            if len(unresolved) < 40:
                unresolved.append({"cd_ref": row["cd_ref"], "taxon": row["name"], "reason": "unmatched"})
            continue
        if realm != "flora":
            stats["excluded_realm"] += 1
            continue
        stats["matched"] += 1
        values[row["value"]] += 1
        key = (row["cd_ref"], row["value"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": row["cd_ref"],
                "region": "HDF",
                "category": "znieff",
                "label": "Déterminante ZNIEFF",
                "value": row["value"],
                "sourceId": source["id"],
                "scope": "partial",
                "scopeLabel": SCOPE_LABEL,
            }
        )

    candidates = stats["matched"] + stats["unmatched"]
    match_rate = stats["matched"] / candidates if candidates else 1.0
    covered_refs = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": 2026,
            "official": True,
            "checkedAt": checked_at,
            "sha256": entry["digest"],
            "landingPage": LANDING_URL,
            "sourceUrl": source["sourceUrl"],
        },
        "replaces": [
            {
                "region": "HDF",
                "category": "znieff",
                "realm": "flora",
                "cdRefs": covered_refs,
            },
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": {
            **dict(stats),
            "matchRate": round(match_rate, 6),
            "rawDetermValues": dict(sorted(entry["raw_codes"].items())),
            "values": dict(sorted(values.items())),
            "unresolvedSample": unresolved,
        },
    }


def build_historical_fauna_if_available(args, input_dir: Path, out_dir: Path) -> None:
    fauna_files = [input_dir / "picardie.ods", input_dir / "npdc.ods"]
    if not all(path.is_file() for path in fauna_files):
        print("ZNIEFF faune historique HDF non présente dans ce dossier : étape ignorée.")
        return

    command = [
        sys.executable,
        str(Path(__file__).with_name("build_znieff_fauna.py")),
        "--taxref",
        str(args.taxref),
        "--input-dir",
        str(input_dir),
        "--out-dir",
        str(out_dir),
        "--checked-at",
        str(args.checked_at),
        "--min-match-rate",
        str(args.min_match_rate),
    ]
    subprocess.run(command, check=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    parsed = parse_all(input_dir)
    wanted_codes = wanted_from_parsed(parsed)
    by_cd_ref = taxref_lookup(Path(args.taxref), wanted_codes)

    total_statuses = 0
    for source in SOURCES:
        package = build_package(source, parsed[source["key"]], by_cd_ref, args.checked_at)
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut ZNIEFF Hauts-de-France produit")
        if len(package["statuses"]) < source["minVolume"]:
            raise SystemExit(
                f"{source['id']}: volume ZNIEFF Hauts-de-France anormalement faible: {len(package['statuses'])}"
            )
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / source["outName"]
        output.write_text(
            json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n",
            encoding="utf-8",
        )
        total_statuses += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")

    print(f"Hauts-de-France ZNIEFF Digitale: {len(SOURCES)} paquets, {total_statuses} statuts")
    build_historical_fauna_if_available(args, input_dir, out_dir)


if __name__ == "__main__":
    main()
