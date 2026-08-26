#!/usr/bin/env python3
"""Listes rouges régionales IRPN Hauts-de-France — multi-groupes faune."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

LANDING_URL = "https://irpn.drealnpdc.fr/"
PRODUCER = "IRPN / partenaires naturalistes / CSRPN Hauts-de-France"
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[a-z]{0,3})$")
NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}
TABLE_NS = NS["table"]

CODE_HEADERS = ("CDNOM", "CD_NOM")
CATEGORY_HEADERS = ("CATEGORIE_HAUTS-DE-FRANCE",)
NAME_HEADERS = ("NOM_SCIENTIFIQUE",)

# Chaque paquet IRPN correspond à un classeur téléchargé par download_lrr.sh.
FILES = {
    "oiseaux": {
        "filename": "oiseaux.xlsx",
        "sheet": "BASE_LRHDF_OIS_N",
        "sha256": "72cd2b60ed47120ab08f722bfdd4256d69bdba424722499dc3062a5e0705879a",
        "sourceUrl": (
            "https://irpn.drealnpdc.fr/wp-content/uploads/2024/04/"
            "LRR_oiseaux_nicheurs_HdF_synthese.xlsx"
        ),
    },
    "papillons": {
        "filename": "papillons.xlsx",
        "sheet": "BASE_LRHDF_PAP_J",
        "sha256": "9cea00c3bd53fcbda1f2cee2fd42312efe0b1fcbf7ddc21e8708c62c643090d7",
        "sourceUrl": (
            "https://irpn.drealnpdc.fr/wp-content/uploads/2024/03/"
            "LRR_papillons-de-jour_synthese_mars-1.xlsx"
        ),
    },
    "orthopteres": {
        "filename": "orthopteres.xlsx",
        "sheet": "BASE_LRHDF_ORTHOP",
        "sha256": "ee2263c415105457826952c87e15e36830f4a4640a9e65b71e5209009ffb6e7b",
        "sourceUrl": (
            "http://www.picardie-nature.org/IMG/xlsx/"
            "lrr_orthopteres-mantodea_phasmida.xlsx"
        ),
    },
    "coccinelles": {
        "filename": "coccinelles.xlsx",
        "sheet": "BASE_LRHDF_COCCINELL",
        "sha256": "6be17093692130b2833e4d54d9e7ec051356c132a6231da6fea567608798f565",
        "sourceUrl": (
            "http://www.picardie-nature.org/IMG/xlsx/"
            "lrr_coccinelles_synthese_septembre.xlsx"
        ),
    },
    "mollusques": {
        "filename": "mollusques.ods",
        "sheet": "BASE_LRHDF_MOL_C",
        "sha256": "bc30243f569fc7e9b1807c8cdc8c653da20458a81e81be1a6768c36a88ae3ee1",
        "sourceUrl": (
            "https://irpn.drealnpdc.fr/wp-content/uploads/2025/02/"
            "LRR_mollusques_HdF_synthese.ods"
        ),
    },
    "poissons": {
        "filename": "poissons.ods",
        "sheet": "BASE_LRHDF_POI_ED",
        "sha256": "a1202ece93abf61268bdfcaa1bdda0a83c05ae8f29f0032955323dc2e6d14df0",
        "sourceUrl": (
            "https://irpn.drealnpdc.fr/wp-content/uploads/2025/09/"
            "LRR_poissons_ecrevisses_eau-douce_HDF.ods"
        ),
    },
}

SOURCES = [
    {
        "key": "oiseaux-nicheurs",
        "file_key": "oiseaux",
        "id": "irpn-hdf-lrr-oiseaux-nicheurs-2024",
        "name": "Liste rouge Oiseaux nicheurs Hauts-de-France",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "papillons-jour",
        "file_key": "papillons",
        "id": "irpn-hdf-lrr-papillons-jour-2024",
        "name": "Liste rouge Papillons de jour Hauts-de-France",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "mollusques",
        "file_key": "mollusques",
        "id": "irpn-hdf-lrr-mollusques-2024",
        "name": "Liste rouge Mollusques continentaux Hauts-de-France",
        "version": "2024",
        "year": 2024,
    },
    {
        "key": "poissons-ecrevisses",
        "file_key": "poissons",
        "id": "irpn-hdf-lrr-poissons-ecrevisses-2025",
        "name": "Liste rouge Poissons et écrevisses d'eau douce Hauts-de-France",
        "version": "2025",
        "year": 2025,
    },
    {
        "key": "orthopteres",
        "file_key": "orthopteres",
        "id": "irpn-hdf-lrr-orthopteres-2025",
        "name": "Liste rouge Orthoptères, mantes et phasmes Hauts-de-France",
        "version": "2025",
        "year": 2025,
    },
    {
        "key": "coccinelles",
        "file_key": "coccinelles",
        "id": "irpn-hdf-lrr-coccinelles-2025",
        "name": "Liste rouge Coccinelles Hauts-de-France",
        "version": "2025",
        "year": 2025,
    },
]


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def normalize_category(value: object) -> str | None:
    text = clean(value)
    match = re.fullmatch(r"(?i)(EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA)([A-Za-z]{0,3})", text)
    if not match:
        return None
    base = match.group(1).upper()
    if base.startswith("CR") and "*" in match.group(1):
        base = "CR*"
    suffix = match.group(2)
    if base == "NA" and suffix:
        category = "NA" + suffix.lower()
    elif suffix:
        return None
    else:
        category = base
    if VALID_LRR_CATEGORY.fullmatch(category):
        return category
    return None


def ods_sheet_rows(path: Path, sheet_name: str) -> list[list[str]]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))
    for table in root.findall(".//table:table", NS):
        if table.get(f"{{{TABLE_NS}}}name") != sheet_name:
            continue
        rows: list[list[str]] = []
        for table_row in table.findall("table:table-row", NS):
            cells: list[str] = []
            for cell in table_row.findall("table:table-cell", NS):
                repeat = int(cell.get(f"{{{TABLE_NS}}}number-columns-repeated") or 1)
                text = clean("".join(cell.itertext()))
                cells.extend([text] * min(repeat, 20))
            if any(cells):
                rows.append(cells)
        return rows
    raise RuntimeError(f"{path.name}: onglet introuvable {sheet_name}")


def xlsx_sheet_rows(path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"{path.name}: onglet introuvable {sheet_name}")
        rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
    finally:
        workbook.close()
    return rows


def sheet_rows(path: Path, sheet_name: str) -> list[list[object]]:
    if path.suffix.lower() == ".ods":
        return ods_sheet_rows(path, sheet_name)
    return xlsx_sheet_rows(path, sheet_name)


def column_index(headers: list[object], candidates: tuple[str, ...]) -> int | None:
    normalized = [normalize(value) for value in headers]
    for candidate in candidates:
        target = normalize(candidate)
        if target in normalized:
            return normalized.index(target)
    return None


def find_header_row(rows: list[list[object]]) -> int:
    for index, headers in enumerate(rows[:10]):
        if (
            column_index(headers, CODE_HEADERS) is not None
            and column_index(headers, CATEGORY_HEADERS) is not None
            and column_index(headers, NAME_HEADERS) is not None
        ):
            return index
    raise RuntimeError("ligne d'en-tête introuvable (CDNOM/CD_NOM, CATEGORIE_HAUTS-DE-FRANCE, NOM_SCIENTIFIQUE)")


def read_file_rows(file_key: str, meta: dict, input_dir: Path) -> list[dict]:
    path = input_dir / meta["filename"]
    rows = sheet_rows(path, meta["sheet"])
    header_index = find_header_row(rows)
    headers = rows[header_index]
    code_index = column_index(headers, CODE_HEADERS)
    category_index = column_index(headers, CATEGORY_HEADERS)
    name_index = column_index(headers, NAME_HEADERS)
    if code_index is None or category_index is None or name_index is None:
        raise RuntimeError(f"{meta['filename']}: colonnes CDNOM/CATEGORIE/NOM_SCIENTIFIQUE introuvables")

    parsed = []
    for values in rows[header_index + 1 :]:
        category = normalize_category(values[category_index] if category_index < len(values) else None)
        if category is None:
            continue
        parsed.append(
            {
                "code": as_int(values[code_index] if code_index < len(values) else None),
                "name": clean(values[name_index] if name_index < len(values) else ""),
                "category": category,
            }
        )
    if not parsed:
        raise RuntimeError(f"{meta['filename']}: aucune ligne de statut exploitable ({file_key})")
    return parsed


def parse_all(input_dir: Path) -> dict[str, list[dict]]:
    for meta in FILES.values():
        path = input_dir / meta["filename"]
        actual = sha256(path)
        if actual != meta["sha256"]:
            raise SystemExit(f"{path.name}: SHA-256 inattendu {actual} != {meta['sha256']}")

    parsed: dict[str, list[dict]] = {}
    for file_key, meta in FILES.items():
        parsed[file_key] = read_file_rows(file_key, meta, input_dir)
    return parsed


def wanted_from_parsed(parsed: dict[str, list[dict]]):
    codes: set[int] = set()
    names: set[str] = set()
    for rows in parsed.values():
        for row in rows:
            if row["code"] is not None:
                codes.add(row["code"])
            if row["name"]:
                names.add(normalize(row["name"]))
    return codes, names


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str]):
    by_cd_nom: dict[int, tuple[int, str | None]] = {}
    by_name = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom_raw = str(row.get("CD_NOM") or "").strip()
            cd_ref_raw = str(row.get("CD_REF") or "").strip()
            if not cd_nom_raw.isdigit() or not cd_ref_raw.isdigit():
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            cd_nom = int(cd_nom_raw)
            cd_ref = int(cd_ref_raw)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = (cd_ref, realm)
            if realm and wanted_names:
                for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                    label = row.get(field)
                    if label and normalize(label) in wanted_names:
                        by_name[normalize(label)].add((cd_ref, realm))
    return by_cd_nom, by_name


def resolve(row, by_cd_nom, by_name):
    code = row["code"]
    if code is not None and code in by_cd_nom:
        cd_ref, realm = by_cd_nom[code]
        if realm:
            return cd_ref, realm, "cd_nom"
        return None, None, "excluded_realm"
    if row["name"]:
        candidates = by_name.get(normalize(row["name"]), set())
        if len(candidates) == 1:
            cd_ref, realm = next(iter(candidates))
            return cd_ref, realm, "name"
        if len(candidates) > 1:
            return None, None, "ambiguous"
    return None, None, "unmatched"


def build_package(source, rows, input_dir: Path, by_cd_nom, by_name, checked_at: str):
    stats = {
        "rows": len(rows),
        "matched": 0,
        "cd_nom": 0,
        "name": 0,
        "excluded_realm": 0,
        "unmatched": 0,
        "ambiguous": 0,
        "unexpectedRealm": 0,
        "unresolvedSample": [],
        "values": {},
    }
    values = defaultdict(int)
    statuses = []
    seen = set()

    for row in rows:
        cd_ref, realm, mode = resolve(row, by_cd_nom, by_name)
        if cd_ref is None or realm is None:
            stats[mode] += 1
            if len(stats["unresolvedSample"]) < 30:
                stats["unresolvedSample"].append(
                    {
                        "code": row["code"],
                        "taxon": row["name"],
                        "category": row["category"],
                        "reason": mode,
                    }
                )
            continue
        if realm != "fauna":
            stats["unexpectedRealm"] += 1
            continue
        stats["matched"] += 1
        stats[mode] += 1
        values[row["category"]] += 1
        key = (cd_ref, row["category"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": cd_ref,
                "region": "HDF",
                "category": "red_list_regional",
                "label": "Liste rouge régionale",
                "value": row["category"],
                "sourceId": source["id"],
                "scope": "regional",
            }
        )

    candidates = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    file_meta = FILES[source["file_key"]]
    file_path = input_dir / file_meta["filename"]
    covered_refs = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": sha256(file_path),
            "landingPage": LANDING_URL,
            "sourceUrl": file_meta["sourceUrl"],
        },
        "replaces": [
            {
                "region": "HDF",
                "category": "red_list_regional",
                "realm": "fauna",
                "cdRefs": covered_refs,
            },
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": stats,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    parsed = parse_all(input_dir)
    wanted_codes, wanted_names = wanted_from_parsed(parsed)
    by_cd_nom, by_name = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names)

    total_statuses = 0
    for source in SOURCES:
        package = build_package(
            source, parsed[source["file_key"]], input_dir, by_cd_nom, by_name, args.checked_at
        )
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        if "url" in package["source"]:
            raise SystemExit(f"{source['id']}: champ url interdit")
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"hdf-lrr-{source['key']}.json"
        output.write_text(
            json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n",
            encoding="utf-8",
        )
        total_statuses += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")

    print(f"Hauts-de-France LRR IRPN: {len(SOURCES)} paquets, {total_statuses} statuts")


if __name__ == "__main__":
    main()
