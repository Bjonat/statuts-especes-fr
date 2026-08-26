#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[A-Z]?)$")

LANDING_URL = (
    "https://www.normandie.developpement-durable.gouv.fr/"
    "les-listes-rouges-dans-le-monde-et-en-normandie-a6663.html"
)

SOURCES = [
    {
        "key": "oiseaux",
        "filename": "oiseaux-2024.xlsx",
        "id": "anbdd-normandie-lrr-oiseaux-nicheurs-2024",
        "name": "Liste rouge Oiseaux nicheurs Normandie",
        "version": "2024",
        "year": 2024,
        "category_headers": ["Catégorie 2024"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/wp-content/uploads/2025/01/LR_Oiseaux_Nicheurs-_Normandie_2024_telechargement.xlsx",
    },
    {
        "key": "mammiferes",
        "filename": "mammiferes-2022.xlsx",
        "id": "anbdd-normandie-lrr-mammiferes-2022",
        "name": "Liste rouge Mammifères Normandie",
        "version": "2022",
        "year": 2022,
        "category_headers": ["Catégorie Liste rouge Normandie"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/wp-content/uploads/2022/11/Tableau_Liste_Rouge_Mammiferes_Normandie_2022-VF.xlsx",
    },
    {
        "key": "amphibiens",
        "filename": "amphibiens-2022.xlsx",
        "id": "anbdd-normandie-lrr-amphibiens-2022",
        "name": "Liste rouge Amphibiens Normandie",
        "version": "2022",
        "year": 2022,
        "category_headers": ["Catégorie Liste rouge Normandie (2022)"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/wp-content/uploads/2022/09/Tableau-LR-Amphibiens-Normandie-2022.xlsx",
    },
    {
        "key": "reptiles",
        "filename": "reptiles-2022.xlsx",
        "id": "anbdd-normandie-lrr-reptiles-2022",
        "name": "Liste rouge Reptiles Normandie",
        "version": "2022",
        "year": 2022,
        "category_headers": ["Catégorie Liste rouge Normandie (2022)"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/publication/liste-rouge-des-reptiles-de-normandie/tableau-lr-reptiles-normandie-2022/",
    },
    {
        "key": "odonates",
        "filename": "odonates.xlsx",
        "id": "anbdd-normandie-lrr-odonates-2022",
        "name": "Liste rouge Odonates Normandie",
        "version": "2022",
        "year": 2022,
        "category_headers": ["LISTE ROUGE NORMANDE"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/wp-content/uploads/2022/09/LR_ODONATES_tableau-de-synthese_pour_ANBDD_VF.xlsx",
    },
    {
        "key": "orthopteres",
        "filename": "orthopteres-2022.xlsx",
        "id": "anbdd-normandie-lrr-orthopteres-2022",
        "name": "Liste rouge Orthoptères, mantes et phasmes Normandie",
        "version": "2022",
        "year": 2022,
        "category_headers": ["LISTE ROUGE NORMANDE"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/wp-content/uploads/2023/05/Liste_Rouge_ORTHO_Normandie_ANBDD_2022.xlsx",
    },
    {
        "key": "rhopaloceres",
        "filename": "rhopaloceres.xlsx",
        "id": "anbdd-normandie-lrr-rhopaloceres-2022",
        "name": "Liste rouge Rhopalocères et zygènes Normandie",
        "version": "2022",
        "year": 2022,
        "category_headers": ["LISTE ROUGE NORMANDE"],
        "landingPage": LANDING_URL,
        "sourceUrl": "https://www.anbdd.fr/wp-content/uploads/2022/05/LR_RHOPALO_tableau-de-synthese_ANBDD.xlsx",
    },
]

CODE_HEADERS = ["CD_NOM", "CD_NOM (V15)", "CD_NOM (V10)"]
NAME_HEADERS = [
    "Nom scientifique",
    "Nom scientifique taxref V15",
    "Nom scientifique taxref V10",
    "Nom scientifique court",
]


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def find_header_row(sheet, source):
    required_categories = {normalize(value) for value in source["category_headers"]}
    for row_index, values in enumerate(sheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        headers = [normalize(value) for value in values]
        has_category = any(header in required_categories for header in headers)
        has_code = any(normalize(candidate) in headers for candidate in CODE_HEADERS)
        if has_category and has_code:
            return row_index, list(values)
    raise RuntimeError(f"{source['filename']}: ligne d'en-tête introuvable")


def column_index(headers, alternatives):
    normalized = [normalize(value) for value in headers]
    for alternative in alternatives:
        target = normalize(alternative)
        if target in normalized:
            return normalized.index(target)
    return None


def read_source_rows(path: Path, source):
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        try:
            header_row, headers = find_header_row(sheet, source)
        except RuntimeError:
            continue
        code_index = column_index(headers, CODE_HEADERS)
        category_index = column_index(headers, source["category_headers"])
        name_index = column_index(headers, NAME_HEADERS)
        if code_index is None or category_index is None:
            continue
        rows = []
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            category = str(values[category_index] or "").strip().upper()
            if not VALID_LRR_CATEGORY.fullmatch(category):
                continue
            code = as_int(values[code_index])
            name = str(values[name_index] or "").strip() if name_index is not None else ""
            rows.append({"code": code, "name": name, "category": category})
        if rows:
            return rows
    raise RuntimeError(f"{source['filename']}: aucune table de statuts exploitable")


def wanted_from_sources(input_dir: Path):
    codes: set[int] = set()
    names: set[str] = set()
    parsed = {}
    for source in SOURCES:
        path = input_dir / source["filename"]
        rows = read_source_rows(path, source)
        parsed[source["key"]] = rows
        for row in rows:
            if row["code"] is not None:
                codes.add(row["code"])
            if row["name"]:
                names.add(normalize(row["name"]))
    return parsed, codes, names


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str]):
    by_cd_nom: dict[int, tuple[int, str | None]] = {}
    by_name = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom_raw = str(row.get("CD_NOM") or "").strip()
            cd_ref_raw = str(row.get("CD_REF") or "").strip()
            if not cd_nom_raw.isdigit() or not cd_ref_raw.isdigit():
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            cd_nom = int(cd_nom_raw)
            cd_ref = int(cd_ref_raw)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = (cd_ref, realm)
            if realm and wanted_names:
                for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                    label = row.get(field)
                    if label and normalize(label) in wanted_names:
                        by_name[normalize(label)].add((cd_ref, realm))
    return by_cd_nom, by_name


def resolve(row, by_cd_nom, by_name):
    code = row["code"]
    if code is not None and code in by_cd_nom:
        cd_ref, realm = by_cd_nom[code]
        if realm:
            return cd_ref, realm, "cd_nom"
        return None, None, "excluded_realm"
    if row["name"]:
        candidates = by_name.get(normalize(row["name"]), set())
        if len(candidates) == 1:
            cd_ref, realm = next(iter(candidates))
            return cd_ref, realm, "name"
        if len(candidates) > 1:
            return None, None, "ambiguous"
    return None, None, "unmatched"


def build_package(source, rows, input_dir: Path, by_cd_nom, by_name, checked_at: str):
    stats = {
        "rows": len(rows),
        "matched": 0,
        "cd_nom": 0,
        "name": 0,
        "excluded_realm": 0,
        "unmatched": 0,
        "ambiguous": 0,
        "unexpectedRealm": 0,
        "unresolvedSample": [],
        "values": {},
    }
    values = defaultdict(int)
    statuses = []
    seen = set()

    for row in rows:
        cd_ref, realm, mode = resolve(row, by_cd_nom, by_name)
        if cd_ref is None or realm is None:
            stats[mode] += 1
            if len(stats["unresolvedSample"]) < 30:
                stats["unresolvedSample"].append({
                    "code": row["code"], "taxon": row["name"], "category": row["category"], "reason": mode
                })
            continue
        if realm != "fauna":
            stats["unexpectedRealm"] += 1
            continue
        stats["matched"] += 1
        stats[mode] += 1
        values[row["category"]] += 1
        key = (cd_ref, row["category"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append({
            "cdRef": cd_ref,
            "region": "NOR",
            "category": "red_list_regional",
            "label": "Liste rouge régionale",
            "value": row["category"],
            "sourceId": source["id"],
            "scope": "regional",
        })

    candidates = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    file_path = input_dir / source["filename"]
    covered_refs = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": "Agence normande de la biodiversité et du développement durable / partenaires naturalistes / CSRPN Normandie",
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "landingPage": source["landingPage"],
            "sourceUrl": source["sourceUrl"],
            "sha256": sha256(file_path),
        },
        "replaces": [
            {"region": "NOR", "category": "red_list_regional", "realm": "fauna", "cdRefs": covered_refs},
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": stats,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    parsed, wanted_codes, wanted_names = wanted_from_sources(input_dir)
    by_cd_nom, by_name = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names)

    total_statuses = 0
    for source in SOURCES:
        package = build_package(source, parsed[source["key"]], input_dir, by_cd_nom, by_name, args.checked_at)
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"nor-lrr-{source['key']}.json"
        output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        total_statuses += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")

    print(f"Normandie: {len(SOURCES)} paquets, {total_statuses} statuts LRR")


if __name__ == "__main__":
    main()
