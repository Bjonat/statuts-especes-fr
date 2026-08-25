#!/usr/bin/env python3
"""ZNIEFF flore Haute-Normandie via catalogue Digitale CBNHDF (portée partielle)."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

SOURCE_ID = "cbnhdf-digitale-znieff-hn-flora-2026-03-31"
SOURCE_URL = "https://www.cbnhdf.fr/system/files/2026-05/DIGITALE_BS-BIF-FVF_PV_4.0_20260331.xlsx"
LANDING_URL = "https://www.cbnhdf.fr/je-telecharge"
DREAL_LANDING = "https://www.normandie.developpement-durable.gouv.fr/les-listes-d-especes-et-d-habitats-determinants-de-a3126.html"
PRODUCER = "Conservatoire botanique national des Hauts-de-France / CSRPN (Haute-Normandie)"
EXPECTED_SHA256 = "71ae71b770f7b3911349e501caaaa65ac7dba8172d12b96ef4b90d5056995c95"
SHEET_NAME = "REG-DIGITALE-BS-BIF-FVF-PV_4.0"
SCOPE_LABEL = "Haute-Normandie"
TERRITORY = "HN"
# [Oui] = erreur / douteux / cultivé uniquement — hors publication déterminante.
VALUE_BY_CODE = {
    "Oui": "Oui",
    "(Oui)": "Oui (disparu/présumé)",
    "pp": "Oui (pro parte)",
}
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
MAX_VALUE_LENGTH = 80


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def read_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise SystemExit(f"Feuille absente: {SHEET_NAME}")
    worksheet = workbook[SHEET_NAME]
    iterator = worksheet.iter_rows(values_only=True)
    header = [clean(value) for value in next(iterator)]
    required = {"CH_Territoire", "CH_DetermZNIEFF", "CD_REF_TAXREF"}
    if not required.issubset(set(header)):
        workbook.close()
        raise SystemExit(f"Colonnes manquantes: {sorted(required - set(header))}")
    rows = []
    for values in iterator:
        row = {header[index]: values[index] if index < len(values) else None for index in range(len(header))}
        rows.append(row)
    workbook.close()
    return rows


def taxref_lookup(path: Path, wanted_codes: set[int]):
    by_cd_ref: dict[int, str | None] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_ref_raw = clean(row.get("CD_REF"))
            if not cd_ref_raw.isdigit():
                continue
            cd_ref = int(cd_ref_raw)
            if cd_ref not in wanted_codes:
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            by_cd_ref[cd_ref] = realm
    return by_cd_ref


def build_package(taxref_path: Path, source_path: Path, checked_at: str):
    digest = sha256(source_path)
    if digest != EXPECTED_SHA256:
        raise SystemExit(f"SHA-256 Digitale inattendu: {digest}")

    rows = read_rows(source_path)
    selected = []
    raw_codes = Counter()
    for row in rows:
        if clean(row.get("CH_Territoire")) != TERRITORY:
            continue
        code = clean(row.get("CH_DetermZNIEFF"))
        raw_codes[code or "<vide>"] += 1
        value = VALUE_BY_CODE.get(code)
        if not value:
            continue
        if len(value) > MAX_VALUE_LENGTH:
            continue
        cd_ref = as_int(row.get("CD_REF_TAXREF"))
        if cd_ref is None:
            continue
        selected.append({
            "cd_ref": cd_ref,
            "value": value,
            "name": clean(row.get("CH_NomCompletTAXREF") or row.get("CH_NomComp")),
            "raw": code,
        })

    wanted = {row["cd_ref"] for row in selected}
    by_cd_ref = taxref_lookup(taxref_path, wanted)

    stats = Counter()
    stats["rows_source"] = len(rows)
    stats["rows_hn_znieff"] = len(selected)
    values = Counter()
    statuses = []
    seen = set()
    unresolved = []

    for row in selected:
        realm = by_cd_ref.get(row["cd_ref"])
        if realm is None:
            stats["unmatched"] += 1
            if len(unresolved) < 40:
                unresolved.append({"cd_ref": row["cd_ref"], "taxon": row["name"], "reason": "unmatched"})
            continue
        if realm != "flora":
            stats["excluded_realm"] += 1
            continue
        stats["matched"] += 1
        values[row["value"]] += 1
        key = (row["cd_ref"], row["value"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append({
            "cdRef": row["cd_ref"],
            "region": "NOR",
            "category": "znieff",
            "label": "Déterminante ZNIEFF",
            "value": row["value"],
            "sourceId": SOURCE_ID,
            "scope": "partial",
            "scopeLabel": SCOPE_LABEL,
        })

    candidates = stats["matched"] + stats["unmatched"]
    match_rate = stats["matched"] / candidates if candidates else 1.0
    covered_refs = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": SOURCE_ID,
            "name": "Espèces déterminantes ZNIEFF flore Haute-Normandie",
            "producer": PRODUCER,
            "version": "Digitale BS-BIF-FVF-PV 4.0 (2026-03-31)",
            "publicationYear": 2026,
            "official": True,
            "checkedAt": checked_at,
            "sha256": digest,
            "landingPage": LANDING_URL,
            "sourceUrl": SOURCE_URL,
            "mirrorLandingPage": DREAL_LANDING,
        },
        "replaces": [
            {
                "region": "NOR",
                "category": "znieff",
                "realm": "flora",
                "cdRefs": covered_refs,
            },
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": {
            **dict(stats),
            "matchRate": round(match_rate, 6),
            "rawDetermValues": dict(sorted(raw_codes.items())),
            "values": dict(sorted(values.items())),
            "unresolvedSample": unresolved,
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--source", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    package = build_package(Path(args.taxref), Path(args.source), args.checked_at)
    diagnostics = package["diagnostics"]
    print(json.dumps(diagnostics, ensure_ascii=False, indent=2))
    if diagnostics["matchRate"] < args.min_match_rate:
        raise SystemExit(
            f"Taux de raccord TAXREF insuffisant: {diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
        )
    if not package["statuses"]:
        raise SystemExit("Aucun statut ZNIEFF Haute-Normandie produit")
    if len(package["statuses"]) < 500:
        raise SystemExit(f"Volume ZNIEFF HN anormalement faible: {len(package['statuses'])}")
    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(f"Paquet régional écrit: {output} - {len(package['statuses'])} statuts")


if __name__ == "__main__":
    main()
