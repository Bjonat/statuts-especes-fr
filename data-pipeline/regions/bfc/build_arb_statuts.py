#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

WITNESS_SOURCE_ID = "arb-bfc-statuts-2023-12-19"
WITNESS_SHA256 = "0912139a6f6b6902d6be22e383471b971782502e155b5ae83526bddacbcac073"
WITNESS_VERSION = "2023-12-19"
WITNESS_YEAR = 2023
WITNESS_URL = "https://www.arb-bfc.fr/content/uploads/2024/06/231219_sp_statuts_bfc_a_diffuser.xlsx"
DREAL_2026_URL = "https://www.bourgogne-franche-comte.developpement-durable.gouv.fr/IMG/xlsx/260303_sp_statuts_bfc.xlsx"
DREAL_LANDING_URL = "https://www.bourgogne-franche-comte.developpement-durable.gouv.fr/statut-des-especes-a10460.html"
ARB_LANDING_URL = "https://www.arb-bfc.fr/"

REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
SEARCHABLE_RANKS = {"ES", "SSES", "VAR", "SVAR", "FO", "CAR", "RACE", "AGES"}
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[A-Z]?)$")
MAX_VALUE_LENGTH = 80

ZNIEFF_VALUES = {
    "determinante stricte": "Oui",
    "determinante station": "Oui",
    "determinante sous conditions": "Oui sous condition",
    "determinante sous conditions geographiques": "Oui sous condition",
    "determinante sous conditions geographiques + station": "Oui sous condition",
    "determinante stricte + sous conditions": "Oui",
}

REQUIRED_COLUMNS = {
    "cd_ref",
    "nom_scientifique",
    "znieff_determinantes_bfc",
    "znieff_conditions",
    "liste_rouge_bourgogne",
    "liste_rouge_franche_comte",
}


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").replace("–", "-").replace("—", "-").replace("‑", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_cd_noms(value: object) -> list[int]:
    codes = []
    for token in re.split(r"[^0-9]+", clean(value)):
        if token.isdigit():
            codes.append(int(token))
    return codes


def znieff_value(raw: str) -> str | None:
    key = normalize(raw)
    if key in ZNIEFF_VALUES:
        return ZNIEFF_VALUES[key]
    if "sous condition" in key:
        return "Oui sous condition"
    if "determinante" in key:
        return "Oui"
    return None


def compact_value(value: str) -> str | None:
    text = clean(value)
    if not text or len(text) > MAX_VALUE_LENGTH:
        return None
    return text


def add_status(statuses: list[dict], seen: set, record: dict) -> bool:
    key = (
        record["cdRef"],
        record["category"],
        record["label"],
        record["value"],
        record["scope"],
        record.get("scopeLabel", ""),
    )
    if key in seen:
        return False
    seen.add(key)
    statuses.append(record)
    return True


def read_source_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Liste_espèces" not in workbook.sheetnames:
        raise RuntimeError("Feuille Liste_espèces absente du tableur BFC")
    rows = list(workbook["Liste_espèces"].iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Feuille Liste_espèces vide")
    header = [clean(value) for value in rows[0]]
    missing = REQUIRED_COLUMNS - set(header)
    if missing:
        raise RuntimeError(f"Colonnes BFC manquantes: {sorted(missing)}")
    records = []
    for values in rows[1:]:
        record = {name: values[index] if index < len(values) else None for index, name in enumerate(header)}
        if as_int(record.get("cd_ref")) or clean(record.get("nom_scientifique")):
            records.append(record)
    return records


def is_relevant(row: dict) -> bool:
    if clean(row.get("znieff_determinantes_bfc")):
        return True
    for field in ("liste_rouge_bourgogne", "liste_rouge_franche_comte"):
        category = clean(row.get(field)).upper()
        if category and VALID_LRR_CATEGORY.fullmatch(category):
            return True
    return False


def collect_wanted(rows: list[dict]) -> tuple[set[int], set[str]]:
    codes: set[int] = set()
    names: set[str] = set()
    for row in rows:
        code = as_int(row.get("cd_ref"))
        if code:
            codes.add(code)
        codes.update(parse_cd_noms(row.get("cd_nom")))
        name = clean(row.get("nom_scientifique"))
        if name:
            names.add(normalize(name))
    return codes, names


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str]):
    by_cd_nom: dict[int, tuple[int, str | None, str]] = {}
    by_cd_ref: dict[int, tuple[int, str | None, str]] = {}
    by_name: dict[str, set[tuple[int, str, str]]] = defaultdict(set)

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom = as_int(row.get("CD_NOM"))
            cd_ref = as_int(row.get("CD_REF"))
            if not cd_nom or not cd_ref:
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            rank = clean(row.get("RANG")).upper()
            entry = (cd_ref, realm, rank)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = entry
            if cd_ref in wanted_codes and cd_nom == cd_ref:
                by_cd_ref[cd_ref] = entry
            if realm:
                for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                    label = clean(row.get(field))
                    key = normalize(label)
                    if label and key in wanted_names:
                        by_name[key].add((cd_ref, realm, rank))
    return by_cd_nom, by_cd_ref, by_name


def resolve(row, by_cd_nom, by_cd_ref, by_name):
    candidates: list[tuple[tuple[int, str | None, str], str]] = []
    code = as_int(row.get("cd_ref"))
    if code and code in by_cd_nom:
        candidates.append((by_cd_nom[code], "cd_ref"))
    elif code and code in by_cd_ref:
        candidates.append((by_cd_ref[code], "cd_ref"))
    else:
        for synonym in parse_cd_noms(row.get("cd_nom")):
            if synonym in by_cd_nom:
                candidates.append((by_cd_nom[synonym], "cd_nom"))
                break

    if candidates:
        (cd_ref, realm, rank), mode = candidates[0]
        if rank not in SEARCHABLE_RANKS:
            return None, None, rank, "excluded_rank"
        if not realm:
            return None, None, rank, "excluded_realm"
        return cd_ref, realm, rank, mode

    name = clean(row.get("nom_scientifique"))
    if name:
        matches = {
            entry for entry in by_name.get(normalize(name), set()) if entry[2] in SEARCHABLE_RANKS and entry[1]
        }
        if len(matches) == 1:
            cd_ref, realm, rank = next(iter(matches))
            return cd_ref, realm, rank, "name"
        if len(matches) > 1:
            return None, None, "", "ambiguous"
    return None, None, "", "unmatched"


def build_package(taxref_path: Path, source_path: Path, checked_at: str, allow_witness: bool):
    digest = sha256(source_path)
    if digest != WITNESS_SHA256:
        raise SystemExit(
            "SHA-256 BFC inattendu. Le millésime DREAL 2026 n’est pas encore validé par cet adaptateur ; "
            f"reçu {digest}, témoin {WITNESS_SHA256}."
        )
    if not allow_witness:
        raise SystemExit(
            "Le tableur ARB 2023-12-19 est un témoin de schéma. "
            "Passez --allow-witness-millesime pour un smoke-test, jamais pour une publication."
        )

    rows = read_source_rows(source_path)
    relevant = [row for row in rows if is_relevant(row)]
    codes, names = collect_wanted(relevant)
    by_cd_nom, by_cd_ref, by_name = taxref_lookup(taxref_path, codes, names)

    diagnostics = Counter()
    diagnostics["rows_source"] = len(rows)
    diagnostics["rows_relevant"] = len(relevant)
    znieff_values = Counter()
    lrr_values = Counter()
    unresolved = []
    statuses: list[dict] = []
    seen: set = set()
    replacement_refs = {"znieff": {"flora": set(), "fauna": set()}, "red_list_regional": {"flora": set(), "fauna": set()}}

    for row in relevant:
        cd_ref, realm, rank, mode = resolve(row, by_cd_nom, by_cd_ref, by_name)
        diagnostics[mode] += 1
        if cd_ref is None or realm is None:
            if mode in {"unmatched", "ambiguous"} and len(unresolved) < 40:
                unresolved.append({
                    "cd_ref": row.get("cd_ref"),
                    "taxon": clean(row.get("nom_scientifique")),
                    "reason": mode,
                })
            continue
        diagnostics["matched"] += 1
        diagnostics[realm] += 1

        znieff_raw = clean(row.get("znieff_determinantes_bfc"))
        if znieff_raw:
            znieff_values[znieff_raw] += 1
            value = znieff_value(znieff_raw)
            if value is None:
                diagnostics["unexpectedZnieff"] += 1
            elif add_status(statuses, seen, {
                "cdRef": cd_ref,
                "region": "BFC",
                "category": "znieff",
                "label": "Déterminante ZNIEFF",
                "value": value,
                "sourceId": WITNESS_SOURCE_ID,
                "scope": "regional",
            }):
                diagnostics["znieffStatuses"] += 1
                replacement_refs["znieff"][realm].add(cd_ref)

            condition = compact_value(row.get("znieff_conditions"))
            if clean(row.get("znieff_conditions")) and condition is None:
                diagnostics["omittedLongConditions"] += 1
            elif condition and add_status(statuses, seen, {
                "cdRef": cd_ref,
                "region": "BFC",
                "category": "znieff",
                "label": "Condition de déterminance ZNIEFF",
                "value": condition,
                "sourceId": WITNESS_SOURCE_ID,
                "scope": "regional",
            }):
                diagnostics["conditionStatuses"] += 1

        for field, scope_label in (
            ("liste_rouge_bourgogne", "ancienne région Bourgogne"),
            ("liste_rouge_franche_comte", "ancienne région Franche-Comté"),
        ):
            category = clean(row.get(field)).upper()
            if not category:
                continue
            if not VALID_LRR_CATEGORY.fullmatch(category):
                diagnostics["ambiguousLrr"] += 1
                continue
            lrr_values[f"{scope_label}:{category}"] += 1
            if add_status(statuses, seen, {
                "cdRef": cd_ref,
                "region": "BFC",
                "category": "red_list_regional",
                "label": "Liste rouge régionale",
                "value": category,
                "sourceId": WITNESS_SOURCE_ID,
                "scope": "partial",
                "scopeLabel": scope_label,
            }):
                diagnostics["lrrStatuses"] += 1
                replacement_refs["red_list_regional"][realm].add(cd_ref)

    candidates = diagnostics["matched"] + diagnostics["unmatched"] + diagnostics["ambiguous"]
    match_rate = diagnostics["matched"] / candidates if candidates else 1.0
    replaces = []
    for category, realms in replacement_refs.items():
        for realm, refs in realms.items():
            if refs:
                replaces.append({
                    "region": "BFC",
                    "category": category,
                    "realm": realm,
                    "cdRefs": sorted(refs),
                })

    return {
        "schemaVersion": 1,
        "source": {
            "id": WITNESS_SOURCE_ID,
            "name": "Statuts des espèces de Bourgogne-Franche-Comté",
            "producer": "DREAL Bourgogne-Franche-Comté / ARB BFC / Sigogne / CSRPN Bourgogne-Franche-Comté",
            "version": WITNESS_VERSION,
            "publicationYear": WITNESS_YEAR,
            "official": True,
            "checkedAt": checked_at,
            "sha256": digest,
            "landingPage": DREAL_LANDING_URL,
            "sourceUrl": WITNESS_URL,
            "mirrorLandingPage": ARB_LANDING_URL,
            "canonicalSourceUrl": DREAL_2026_URL,
            "publicationPolicy": "schema-witness-smoke-only",
        },
        "replaces": replaces,
        "statuses": sorted(statuses, key=lambda status: (
            status["cdRef"], status["category"], status["label"], status.get("scopeLabel", ""), status["value"]
        )),
        "diagnostics": {
            **dict(diagnostics),
            "matchRate": round(match_rate, 6),
            "znieffValues": dict(sorted(znieff_values.items())),
            "lrrValues": dict(sorted(lrr_values.items())),
            "unresolvedSample": unresolved,
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--source", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    parser.add_argument(
        "--allow-witness-millesime",
        action="store_true",
        help="Autorise le millésime ARB 2023-12-19 uniquement pour un smoke-test de schéma.",
    )
    args = parser.parse_args()

    package = build_package(
        Path(args.taxref),
        Path(args.source),
        args.checked_at,
        args.allow_witness_millesime,
    )
    diagnostics = package["diagnostics"]
    print(json.dumps(diagnostics, ensure_ascii=False, indent=2))
    if diagnostics["matchRate"] < args.min_match_rate:
        raise SystemExit(
            f"Taux de raccord TAXREF BFC insuffisant: {diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
        )
    if diagnostics.get("znieffStatuses", 0) < 800:
        raise SystemExit(f"Volume ZNIEFF BFC anormalement faible: {diagnostics.get('znieffStatuses', 0)}")
    if diagnostics.get("lrrStatuses", 0) < 4_000:
        raise SystemExit(f"Volume LRR BFC anormalement faible: {diagnostics.get('lrrStatuses', 0)}")
    if not package["statuses"]:
        raise SystemExit("Aucun statut BFC produit")

    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(f"Paquet régional témoin écrit: {output} - {len(package['statuses'])} statuts")
    print("Ce paquet ne doit pas être fusionné dans le dataset officiel tant que le millésime DREAL 2026 n’est pas validé.")


if __name__ == "__main__":
    main()
