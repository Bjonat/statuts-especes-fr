#!/usr/bin/env python3
"""Listes rouges régionales unifiées Grand Est (ODONAT) — multi-groupes faune."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LANDING_URL = "https://www.grand-est.developpement-durable.gouv.fr/listes-rouges-grand-est-a22124.html"
ODONAT_BASE = "https://www.odonat-grandest.fr/telechargements/Listes_rouges"
PRODUCER = "DREAL Grand Est / CSRPN Grand Est / ODONAT Grand Est / partenaires naturalistes"
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[A-Z]{0,3})$")
CODE_HEADERS = ("CD_NOM", "ID_TAX")
CATEGORY_HEADERS = ("Catégorie UICN (détaillée)", "CAT_UICN")
NAME_HEADERS = ("NOM SCIENTIFIQUE", "NOM_SCI_LR")
# Rangs TAXREF / LRGE importables ; les formes populationnelles sans ID propre sont omises.
IMPORTABLE_RANKS = {"ES", "SSES", "VAR", "SVAR", "FO", "CAR", "RACE", "AGES"}

SOURCES = [
    {
        "key": "amphibiens",
        "filename": "amphibiens-reptiles.xlsx",
        "remote": "LISTE_ROUGE_AMPHIBIA_REPTILIA.xlsx",
        "sheet": "LISTE AMPHIBIENS COMPLETE",
        "id": "dreal-ges-odonat-lrr-amphibiens-2023",
        "name": "Liste rouge Amphibiens Grand Est",
        "version": "LRGE v1.0 - septembre 2023",
        "year": 2023,
        "sha256": "0c284571af56c70c31b956994fd0b880c62a1088a267bae6a8eb048f2cbb7d40",
    },
    {
        "key": "reptiles",
        "filename": "amphibiens-reptiles.xlsx",
        "remote": "LISTE_ROUGE_AMPHIBIA_REPTILIA.xlsx",
        "sheet": "LISTE REPTILES COMPLETE",
        "id": "dreal-ges-odonat-lrr-reptiles-2023",
        "name": "Liste rouge Reptiles Grand Est",
        "version": "LRGE v1.0 - septembre 2023",
        "year": 2023,
        "sha256": "0c284571af56c70c31b956994fd0b880c62a1088a267bae6a8eb048f2cbb7d40",
    },
    {
        "key": "mollusques",
        "filename": "mollusques.xlsx",
        "remote": "LISTE_ROUGE_MOLLUSQUES.xlsx",
        "sheet": "LISTE grtax COMPLETE",
        "id": "dreal-ges-odonat-lrr-mollusques-2023-v1.1",
        "name": "Liste rouge Mollusques continentaux Grand Est",
        "version": "LRGE v1.1 - septembre 2023",
        "year": 2023,
        "sha256": "4f87fbaf6a4d9b0541ba73db1a24331df40e41aeae347cdc0cf32441c7512a16",
    },
    {
        "key": "odonates",
        "filename": "odonates.xlsx",
        "remote": "LISTE_ROUGE_ODONATES.xlsx",
        "sheet": "LISTE grtax COMPLETE",
        "id": "dreal-ges-odonat-lrr-odonates-2023",
        "name": "Liste rouge Odonates Grand Est",
        "version": "LRGE v1.0 - septembre 2023",
        "year": 2023,
        "sha256": "9756dfc768502c78526bfa13cf5d7e4c25fbb36a66e2de2ca166fba754e8604a",
    },
    {
        "key": "orthopteres",
        "filename": "orthopteres.xlsx",
        "remote": "LISTE_ROUGE_ORTHOPTERES_v1_0.xlsx",
        "sheet": "LISTE ROUGE ORTHOPTERES compl",
        "id": "dreal-ges-odonat-lrr-orthopteres-2024",
        "name": "Liste rouge Orthoptères et Mantoptères Grand Est",
        "version": "LRGE v1.0 - mai 2024",
        "year": 2024,
        "sha256": "a7737aa5fd8c8a782e7f5b8e3967912705c4b87bb4daff1a9a6b60313f3bee71",
    },
    {
        "key": "oiseaux-nicheurs",
        "filename": "oiseaux-nicheurs.xlsx",
        "remote": "LISTE_ROUGE_OISEAUX_NICHEURS_v1.0.xlsx",
        "sheet": "LISTE ROUGE OISEAUX N détaillée",
        "id": "dreal-ges-odonat-lrr-oiseaux-nicheurs-2024",
        "name": "Liste rouge Oiseaux nicheurs Grand Est",
        "version": "LRGE_OIn_1.0 - 2024",
        "year": 2024,
        "sha256": "e1ae86e45082c11bba9c7b22392cb8adc7efd10697e01b0eab22ce95b09fb07c",
    },
    {
        "key": "branchiopodes",
        "filename": "branchiopodes.xlsx",
        "remote": "LISTE_ROUGE_BRANCHIOPODES.xlsx",
        "sheet": "LISTE ROUGE BRANCHIOPODES cplt",
        "id": "dreal-ges-odonat-lrr-branchiopodes-2025",
        "name": "Liste rouge Grands Branchiopodes Grand Est",
        "version": "LRGE_CRBR_1.0 - 2025",
        "year": 2025,
        "sha256": "0839fe5311c37d1f7557d9890086ac9c83c401f8cfe5362869c836856a50f5e3",
    },
    {
        "key": "decapodes",
        "filename": "decapodes.xlsx",
        "remote": "LISTE_ROUGE_DECAPODES.xlsx",
        "sheet": "LISTE ROUGE DECAPODES détaillée",
        "id": "dreal-ges-odonat-lrr-decapodes-2025",
        "name": "Liste rouge Écrevisses, crabes et crevettes Grand Est",
        "version": "LRGE_ECC_1.0 - 2025",
        "year": 2025,
        "sha256": "7405bb352198c20f0103662dbeb2084fa1890d72ff8dac454386a414205deefe",
    },
    {
        "key": "papillons",
        "filename": "papillons.xlsx",
        "remote": "LISTE_ROUGE_PAPILLONS_JOUR.xlsx",
        "sheet": "LISTE ROUGE PAPILLONS détaillée",
        "id": "dreal-ges-odonat-lrr-papillons-jour-2025",
        "name": "Liste rouge Papillons de jour Grand Est",
        "version": "LRGE_RHO_1.0 - 2025",
        "year": 2025,
        "sha256": "42ed23a590156e41fb272f429803b4cd2ca13076da6fa2f8dc03592ce2640859",
    },
    {
        "key": "poissons",
        "filename": "poissons.xlsx",
        "remote": "LISTE_ROUGE_POISSONS.xlsx",
        "sheet": "LISTE ROUGE POISSONS détaillée",
        "id": "dreal-ges-odonat-lrr-poissons-2024",
        "name": "Liste rouge Poissons d'eau douce Grand Est",
        "version": "LRGE_POI_1.0 - 2024",
        "year": 2024,
        "sha256": "326f5846535a4e12abcbc063a9398650280bb98ed9ff4e048b7705991778d5c4",
    },
]


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def column_index(headers, alternatives):
    normalized = [normalize(value) for value in headers]
    for alternative in alternatives:
        target = normalize(alternative)
        if target in normalized:
            return normalized.index(target)
    return None


def find_header_row(sheet):
    for row_index, values in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        headers = list(values)
        has_code = column_index(headers, CODE_HEADERS) is not None
        has_category = column_index(headers, CATEGORY_HEADERS) is not None
        if has_code and has_category:
            return row_index, headers
    raise RuntimeError("ligne d'en-tête introuvable")


def read_source_rows(path: Path, source: dict) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = source.get("sheet")
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"{source['filename']}: feuille absente ({sheet_name})")
        sheet = workbook[sheet_name]
        header_row, headers = find_header_row(sheet)
        code_index = column_index(headers, CODE_HEADERS)
        category_index = column_index(headers, CATEGORY_HEADERS)
        name_index = column_index(headers, NAME_HEADERS)
        rank_index = column_index(headers, ("RANG", "RANG_LR"))
        if code_index is None or category_index is None:
            raise RuntimeError(f"{source['filename']}: colonnes CD_NOM/CAT introuvables")
        rows = []
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            category = str(values[category_index] or "").strip().upper()
            if not VALID_LRR_CATEGORY.fullmatch(category):
                continue
            code = as_int(values[code_index])
            # Certaines lignes Type B portent ID_TAX = "NA" (hors TAXREF ou forme pop.) :
            # on n'importe que les taxons ancrés par un identifiant numérique propre.
            if code is None:
                continue
            if rank_index is not None:
                rank = str(values[rank_index] or "").strip().upper()
                if rank and rank not in IMPORTABLE_RANKS:
                    continue
            name = str(values[name_index] or "").strip() if name_index is not None else ""
            name = re.sub(r"\s+", " ", name.replace("\n", " ")).strip()
            rows.append({"code": code, "name": name, "category": category})
        if not rows:
            raise RuntimeError(f"{source['id']}: aucune ligne de statut exploitable")
        return rows
    finally:
        workbook.close()


def wanted_from_sources(input_dir: Path):
    codes: set[int] = set()
    names: set[str] = set()
    parsed = {}
    for source in SOURCES:
        path = input_dir / source["filename"]
        digest = sha256(path)
        if digest != source["sha256"]:
            raise RuntimeError(f"{source['id']}: SHA-256 inattendu {digest}")
        rows = read_source_rows(path, source)
        parsed[source["key"]] = rows
        for row in rows:
            if row["code"] is not None:
                codes.add(row["code"])
            if row["name"]:
                names.add(normalize(row["name"]))
    return parsed, codes, names


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str]):
    by_cd_nom: dict[int, tuple[int, str | None]] = {}
    by_name = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom_raw = str(row.get("CD_NOM") or "").strip()
            cd_ref_raw = str(row.get("CD_REF") or "").strip()
            if not cd_nom_raw.isdigit() or not cd_ref_raw.isdigit():
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            cd_nom = int(cd_nom_raw)
            cd_ref = int(cd_ref_raw)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = (cd_ref, realm)
            if realm and wanted_names:
                for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                    label = row.get(field)
                    if label and normalize(label) in wanted_names:
                        by_name[normalize(label)].add((cd_ref, realm))
    return by_cd_nom, by_name


def resolve(row, by_cd_nom, by_name):
    code = row["code"]
    if code is not None and code in by_cd_nom:
        cd_ref, realm = by_cd_nom[code]
        if realm:
            return cd_ref, realm, "cd_nom"
        return None, None, "excluded_realm"
    if row["name"]:
        candidates = by_name.get(normalize(row["name"]), set())
        if len(candidates) == 1:
            cd_ref, realm = next(iter(candidates))
            return cd_ref, realm, "name"
        if len(candidates) > 1:
            return None, None, "ambiguous"
    return None, None, "unmatched"


def build_package(source, rows, input_dir: Path, by_cd_nom, by_name, checked_at: str):
    stats = {
        "rows": len(rows),
        "matched": 0,
        "cd_nom": 0,
        "name": 0,
        "excluded_realm": 0,
        "unmatched": 0,
        "ambiguous": 0,
        "unexpectedRealm": 0,
        "unresolvedSample": [],
        "values": {},
    }
    values = defaultdict(int)
    statuses = []
    seen = set()

    for row in rows:
        cd_ref, realm, mode = resolve(row, by_cd_nom, by_name)
        if cd_ref is None or realm is None:
            stats[mode] += 1
            if len(stats["unresolvedSample"]) < 30:
                stats["unresolvedSample"].append({
                    "code": row["code"],
                    "taxon": row["name"],
                    "category": row["category"],
                    "reason": mode,
                })
            continue
        if realm != "fauna":
            stats["unexpectedRealm"] += 1
            continue
        stats["matched"] += 1
        stats[mode] += 1
        values[row["category"]] += 1
        key = (cd_ref, row["category"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append({
            "cdRef": cd_ref,
            "region": "GES",
            "category": "red_list_regional",
            "label": "Liste rouge régionale",
            "value": row["category"],
            "sourceId": source["id"],
            "scope": "regional",
        })

    candidates = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    file_path = input_dir / source["filename"]
    covered_refs = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": sha256(file_path),
            "landingPage": LANDING_URL,
            "sourceUrl": f"{ODONAT_BASE}/{source['remote']}",
        },
        "replaces": [
            {"region": "GES", "category": "red_list_regional", "realm": "fauna", "cdRefs": covered_refs},
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": stats,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    parsed, wanted_codes, wanted_names = wanted_from_sources(input_dir)
    by_cd_nom, by_name = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names)

    total_statuses = 0
    for source in SOURCES:
        package = build_package(source, parsed[source["key"]], input_dir, by_cd_nom, by_name, args.checked_at)
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"ges-lrr-{source['key']}.json"
        output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        total_statuses += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")

    print(f"Grand Est LRR: {len(SOURCES)} paquets, {total_statuses} statuts")


if __name__ == "__main__":
    main()
