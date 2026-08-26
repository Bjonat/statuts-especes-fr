#!/usr/bin/env python3
"""LRR historiques Grand Est — Alsace mammifères 2014 + Champagne-Ardenne flore 2018."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LANDING_URL = "https://www.grand-est.developpement-durable.gouv.fr/listes-rouges-grand-est-a22124.html"
PRODUCER = "DREAL Grand Est / CSRPN / partenaires (listes d'anciennes régions)"
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[a-z]{0,3})$")

SOURCES = [
    {
        "key": "alsace-mammiferes",
        "filename": "alsace-mammiferes-2014.xlsx",
        "sheet": "Liste_mammiferes",
        "id": "dreal-ges-hist-lrr-alsace-mammiferes-2014",
        "name": "Liste rouge Mammifères Alsace",
        "version": "2014",
        "year": 2014,
        "realm": "fauna",
        "scopeLabel": "Alsace",
        "sha256": "be6cbc009ae804b27d360b63d009dd37c7eaf8b416a5316e4978f76886f21e90",
        "sourceUrl": (
            "https://www.grand-est.developpement-durable.gouv.fr/IMG/xlsx/"
            "liste_rouge_alsace_mammiferes_2014_tableau.xlsx"
        ),
        "code_headers": ("CD_NOM",),
        "category_headers": ("Catégorie Liste rouge Alsace",),
        "name_headers": ("Nom latin", "Nom complet"),
    },
    {
        "key": "ca-flore",
        "filename": "ca-flore-2018.xlsx",
        "sheet": "LISTE_ALPHABETIQUE",
        "id": "dreal-ges-hist-lrr-ca-flore-2018",
        "name": "Liste rouge Flore Champagne-Ardenne",
        "version": "2018",
        "year": 2018,
        "realm": "flora",
        "scopeLabel": "Champagne-Ardenne",
        "sha256": "2017e3dbc1650223f0b6ff847178f2dd9a2c7d49b328369ad297801e62f4818c",
        "sourceUrl": (
            "https://www.grand-est.developpement-durable.gouv.fr/IMG/xlsx/"
            "liste_rouge_champagne_ardenne_flore_2018_validee_uicn.xlsx"
        ),
        "code_headers": ("Cdref", "CD_REF", "CdRef"),
        "category_headers": ("STATUT  FINAL", "STATUT FINAL"),
        "name_headers": ("ESPECE",),
        "code_is_cd_ref": True,
    },
]


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.replace("×", "x").casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def normalize_category(value: object) -> str | None:
    text = clean(value)
    match = re.fullmatch(r"(?i)(EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA)([A-Za-z]{0,3})", text)
    if not match:
        return None
    base = match.group(1).upper()
    if base.startswith("CR") and "*" in match.group(1):
        base = "CR*"
    suffix = match.group(2)
    if base == "NA" and suffix:
        category = "NA" + suffix.lower()
    elif suffix:
        return None
    else:
        category = base
    return category if VALID_LRR_CATEGORY.fullmatch(category) else None


def column_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    normalized = [normalize(header) for header in headers]
    for candidate in candidates:
        target = normalize(candidate)
        if target in normalized:
            return normalized.index(target)
    return None


def read_rows(path: Path, source: dict) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if source["sheet"] not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(f"{path.name}: onglet absent {source['sheet']}")
    rows = list(workbook[source["sheet"]].iter_rows(values_only=True))
    workbook.close()
    headers = [clean(value) for value in rows[0]]
    code_index = column_index(headers, source["code_headers"])
    category_index = column_index(headers, source["category_headers"])
    name_index = column_index(headers, source["name_headers"])
    if code_index is None or category_index is None:
        raise RuntimeError(f"{path.name}: colonnes code/catégorie introuvables ({headers[:12]})")
    parsed = []
    for values in rows[1:]:
        category = normalize_category(values[category_index] if category_index < len(values) else "")
        if category is None:
            continue
        code = as_int(values[code_index] if code_index < len(values) else None)
        name = clean(values[name_index]) if name_index is not None and name_index < len(values) else ""
        parsed.append({"code": code, "name": name, "category": category})
    return parsed


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str], code_is_cd_ref: bool):
    by_code: dict[int, tuple[int, str | None]] = {}
    by_name = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom_raw = clean(row.get("CD_NOM"))
            cd_ref_raw = clean(row.get("CD_REF"))
            if not cd_nom_raw.isdigit() or not cd_ref_raw.isdigit():
                continue
            realm = REALM_BY_KINGDOM.get(normalize(row.get("REGNE")))
            cd_nom = int(cd_nom_raw)
            cd_ref = int(cd_ref_raw)
            key = cd_ref if code_is_cd_ref else cd_nom
            if key in wanted_codes:
                by_code[key] = (cd_ref, realm)
            if realm and wanted_names:
                for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                    label = row.get(field)
                    if label and normalize(label) in wanted_names:
                        by_name[normalize(label)].add((cd_ref, realm))
    return by_code, by_name


def resolve(row, by_code, by_name):
    if row["code"] is not None and row["code"] in by_code:
        cd_ref, realm = by_code[row["code"]]
        if realm:
            return cd_ref, realm, "code"
        return None, None, "excluded_realm"
    if row["name"]:
        candidates = by_name.get(normalize(row["name"]), set())
        if len(candidates) == 1:
            cd_ref, realm = next(iter(candidates))
            return cd_ref, realm, "name"
        if len(candidates) > 1:
            return None, None, "ambiguous"
    return None, None, "unmatched"


def build_package(source, rows, input_dir: Path, by_code, by_name, checked_at: str):
    stats = {
        "rows": len(rows),
        "matched": 0,
        "code": 0,
        "name": 0,
        "unmatched": 0,
        "ambiguous": 0,
        "excluded_realm": 0,
        "unexpectedRealm": 0,
        "unresolvedSample": [],
        "values": {},
    }
    values = defaultdict(int)
    statuses = []
    seen = set()
    for row in rows:
        cd_ref, realm, mode = resolve(row, by_code, by_name)
        if cd_ref is None or realm is None:
            stats[mode] += 1
            if len(stats["unresolvedSample"]) < 20:
                stats["unresolvedSample"].append(
                    {"code": row["code"], "taxon": row["name"], "category": row["category"], "reason": mode}
                )
            continue
        if realm != source["realm"]:
            stats["unexpectedRealm"] += 1
            continue
        stats["matched"] += 1
        stats[mode] += 1
        values[row["category"]] += 1
        key = (cd_ref, row["category"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": cd_ref,
                "region": "GES",
                "category": "red_list_regional",
                "label": "Liste rouge régionale",
                "value": row["category"],
                "sourceId": source["id"],
                "scope": "partial",
                "scopeLabel": source["scopeLabel"],
            }
        )
    candidates = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    covered = sorted({status["cdRef"] for status in statuses})
    path = input_dir / source["filename"]
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": sha256(path),
            "landingPage": LANDING_URL,
            "sourceUrl": source["sourceUrl"],
        },
        "replaces": [
            {
                "region": "GES",
                "category": "red_list_regional",
                "realm": source["realm"],
                "cdRefs": covered,
            },
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": stats,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_dir = Path(args.out_dir)
    total = 0
    for source in SOURCES:
        path = input_dir / source["filename"]
        actual = sha256(path)
        if actual != source["sha256"]:
            raise SystemExit(f"{path.name}: SHA-256 inattendu {actual}")
        rows = read_rows(path, source)
        codes = {row["code"] for row in rows if row["code"] is not None}
        names = {normalize(row["name"]) for row in rows if row["name"]}
        by_code, by_name = taxref_lookup(
            Path(args.taxref), codes, names, source.get("code_is_cd_ref", False)
        )
        package = build_package(source, rows, input_dir, by_code, by_name, args.checked_at)
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord insuffisant {diagnostics['matchRate']:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut")
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"ges-hist-lrr-{source['key']}.json"
        output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        total += len(package["statuses"])
        print(f"Paquet écrit: {output} - {len(package['statuses'])} statuts")
    print(f"GES LRR historiques: {len(SOURCES)} paquets, {total} statuts")


if __name__ == "__main__":
    main()
