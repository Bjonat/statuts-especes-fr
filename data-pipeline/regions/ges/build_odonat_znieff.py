#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

SOURCE_ID = "dreal-ges-odonat-znieff-fauna-2026-v2.2"
SOURCE_URL = "https://www.odonat-grandest.fr/wp-content/uploads/2026/08/listes_especes-determinantes-znieff_grand-est_juin2026.xlsx"
SOURCE_SHA256 = "a130de0436237ebde5b4fcee582ac8890bdadeed679cda96d4896f9082866c1f"
DREAL_URL = "https://www.grand-est.developpement-durable.gouv.fr/IMG/xlsx/listes_edz_aee_faunev2_2_juin2026.xlsx"
LANDING_URL = "https://www.grand-est.developpement-durable.gouv.fr/les-nouvelles-listes-d-especes-determinantes-a22851.html"
ODONAT_LANDING_URL = "https://www.odonat-grandest.fr/znieff-documents-telechargeables/"

SEARCHABLE_RANKS = {"ES", "SSES", "VAR", "SVAR", "FO", "CAR", "RACE", "AGES"}
REGIONAL_CATEGORIES = {"EDZ", "EDZ*", "AEE", "AEE*", "NC", "NE"}
DETERMINANT_CATEGORIES = {"EDZ", "EDZ*"}
DEMOTED_CATEGORIES = {"AEE", "AEE*"}
MAX_VALUE_LENGTH = 80

NATURAL_UNITS = {
    "DETZ_BC": "Plaine de Champagne et Brie",
    "DETZ_AL": "Plateaux lorrains et massif ardennais",
    "DETZ_V": "Massif vosgien",
    "DETZ_RJ": "Fossé rhénan et massif jurassien",
}
PRIORITY_LABELS = {
    "1": "1 - très prioritaire",
    "2": "2 - prioritaire",
    "3": "3 - accompagnatrice",
}


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").replace("–", "-").replace("—", "-").replace("‑", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def workbook_rows(path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Feuille absente: {sheet_name}")
    worksheet = workbook[sheet_name]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def find_header(rows: list[list[object]], required: set[str]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:30]):
        mapping = {clean(value).upper(): column for column, value in enumerate(row) if clean(value)}
        if required.issubset(mapping):
            return index, mapping
    raise RuntimeError(f"En-tête introuvable, champs requis: {sorted(required)}")


def source_rows(path: Path):
    harmonized = workbook_rows(path, "LISTE FAUNE EDZ AEE GRAND EST")
    h_index, h_cols = find_header(harmonized, {"CD_REF", "LB_NOM", "RANG", "DETZ_GE", "DETZ_BC", "DETZ_AL", "DETZ_V", "DETZ_RJ"})

    waiting = workbook_rows(path, "LISTE FAUNE BDD ZNIEFF GRANDEST")
    w_index, w_cols = find_header(waiting, {"CD_REF", "LB_NOM", "RANG", "EDZ_AEE"})

    def records(rows, start, cols):
        for row in rows[start + 1 :]:
            record = {name: row[column] if column < len(row) else None for name, column in cols.items()}
            if any(clean(value) for value in record.values()):
                yield record

    return list(records(harmonized, h_index, h_cols)), list(records(waiting, w_index, w_cols))


def collect_wanted(harmonized, waiting):
    codes: set[int] = set()
    names: set[str] = set()
    for row in [*harmonized, *waiting]:
        code = as_int(row.get("CD_REF"))
        if code:
            codes.add(code)
        name = clean(row.get("LB_NOM"))
        if name:
            names.add(normalize(name))
    return codes, names


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str]):
    by_cd_nom: dict[int, tuple[int, str, str]] = {}
    by_cd_ref: dict[int, tuple[int, str, str]] = {}
    by_name: dict[str, set[tuple[int, str, str]]] = defaultdict(set)

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom = as_int(row.get("CD_NOM"))
            cd_ref = as_int(row.get("CD_REF"))
            if not cd_nom or not cd_ref:
                continue
            kingdom = clean(row.get("REGNE"))
            if normalize(kingdom) != "animalia":
                continue
            rank = clean(row.get("RANG")).upper()
            entry = (cd_ref, "fauna", rank)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = entry
            if cd_ref in wanted_codes and cd_nom == cd_ref:
                by_cd_ref[cd_ref] = entry
            for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                label = clean(row.get(field))
                key = normalize(label)
                if label and key in wanted_names:
                    by_name[key].add(entry)
    return by_cd_nom, by_cd_ref, by_name


def resolve(row, by_cd_nom, by_cd_ref, by_name):
    code = as_int(row.get("CD_REF"))
    candidates = []
    if code and code in by_cd_nom:
        candidates.append((by_cd_nom[code], "cd_nom"))
    elif code and code in by_cd_ref:
        candidates.append((by_cd_ref[code], "cd_ref"))

    if candidates:
        (cd_ref, realm, rank), mode = candidates[0]
        if rank not in SEARCHABLE_RANKS:
            return None, None, rank, "excluded_rank"
        return cd_ref, realm, rank, mode

    name = clean(row.get("LB_NOM"))
    if name:
        matches = {entry for entry in by_name.get(normalize(name), set()) if entry[2] in SEARCHABLE_RANKS}
        if len(matches) == 1:
            cd_ref, realm, rank = next(iter(matches))
            return cd_ref, realm, rank, "name"
        if len(matches) > 1:
            return None, None, "", "ambiguous"
    return None, None, "", "unmatched"


def determinant_value(category: str) -> str:
    if category == "EDZ*":
        return "Oui si (re)découverte"
    return "Oui"


def priority_value(value: object) -> str | None:
    raw = clean(value)
    if not raw:
        return None
    normalized = raw.replace(",", ".")
    if normalized.endswith(".0"):
        normalized = normalized[:-2]
    if normalized in PRIORITY_LABELS:
        return PRIORITY_LABELS[normalized]
    if normalized in {"1*", "2*", "3*"}:
        return PRIORITY_LABELS[normalized[0]] + " *"
    return raw


def compact_value(value: object) -> str | None:
    text = clean(value)
    if not text or len(text) > MAX_VALUE_LENGTH:
        return None
    return text


def add_status(statuses, seen, record):
    key = (
        record["cdRef"], record["category"], record["label"], record["value"],
        record["scope"], record.get("scopeLabel", ""),
    )
    if key in seen:
        return False
    seen.add(key)
    statuses.append(record)
    return True


def build_package(taxref_path: Path, source_path: Path, checked_at: str):
    digest = sha256(source_path)
    if digest != SOURCE_SHA256:
        raise RuntimeError(f"SHA-256 Grand Est inattendu: {digest} != {SOURCE_SHA256}")

    harmonized, waiting = source_rows(source_path)
    wanted_codes, wanted_names = collect_wanted(harmonized, waiting)
    by_cd_nom, by_cd_ref, by_name = taxref_lookup(taxref_path, wanted_codes, wanted_names)

    diagnostics = Counter()
    diagnostics["harmonizedRows"] = len(harmonized)
    diagnostics["waitingRows"] = len(waiting)
    category_counts = Counter()
    waiting_counts = Counter()
    priority_counts = Counter()
    resolution_modes = Counter()
    unresolved = []
    replacement_refs: set[int] = set()
    statuses = []
    seen = set()

    for row in harmonized:
        category = clean(row.get("DETZ_GE")).upper()
        if not category:
            continue
        category_counts[category] += 1
        cd_ref, realm, rank, mode = resolve(row, by_cd_nom, by_cd_ref, by_name)
        resolution_modes[mode] += 1
        if cd_ref is None or realm is None:
            diagnostics[mode] += 1
            if mode not in {"excluded_rank"} and len(unresolved) < 80:
                unresolved.append({
                    "sourceCdRef": as_int(row.get("CD_REF")),
                    "taxon": clean(row.get("LB_NOM")),
                    "rank": clean(row.get("RANG")),
                    "category": category,
                    "reason": mode,
                })
            continue
        diagnostics["matchedHarmonized"] += 1

        if category in REGIONAL_CATEGORIES:
            replacement_refs.add(cd_ref)
        if category not in DETERMINANT_CATEGORIES:
            continue

        if add_status(statuses, seen, {
            "cdRef": cd_ref,
            "region": "GES",
            "category": "znieff",
            "label": "Déterminante ZNIEFF",
            "value": determinant_value(category),
            "sourceId": SOURCE_ID,
            "scope": "regional",
        }):
            diagnostics["regionalDeterminantStatuses"] += 1

        for field, scope_label in NATURAL_UNITS.items():
            value = priority_value(row.get(field))
            if not value:
                continue
            priority_counts[f"{field}:{value}"] += 1
            if add_status(statuses, seen, {
                "cdRef": cd_ref,
                "region": "GES",
                "category": "znieff",
                "label": "Priorité ZNIEFF",
                "value": value,
                "sourceId": SOURCE_ID,
                "scope": "partial",
                "scopeLabel": scope_label,
            }):
                diagnostics["priorityStatuses"] += 1

        condition = compact_value(row.get("DETZ_COND"))
        if clean(row.get("DETZ_COND")) and condition is None:
            diagnostics["omittedLongConditions"] += 1
        elif condition and add_status(statuses, seen, {
            "cdRef": cd_ref,
            "region": "GES",
            "category": "znieff",
            "label": "Condition de déterminance ZNIEFF",
            "value": condition,
            "sourceId": SOURCE_ID,
            "scope": "regional",
        }):
            diagnostics["conditionStatuses"] += 1

        surcot = compact_value(row.get("DETZ_SURCO"))
        if clean(row.get("DETZ_SURCO")) and surcot is None:
            diagnostics["omittedLongSurcotations"] += 1
        elif surcot and add_status(statuses, seen, {
            "cdRef": cd_ref,
            "region": "GES",
            "category": "znieff",
            "label": "Condition de surcotation ZNIEFF",
            "value": surcot,
            "sourceId": SOURCE_ID,
            "scope": "regional",
        }):
            diagnostics["surcotationStatuses"] += 1

    # La feuille BDD rassemble les groupes non encore harmonisés. Lorsqu'un taxon y
    # est explicitement rétrogradé AEE/AEE*, on retire les anciens statuts EDZ des
    # trois ex-régions sans le republier comme espèce déterminante.
    for row in waiting:
        category = clean(row.get("EDZ_AEE")).upper()
        if category:
            waiting_counts[category] += 1
        if category not in DEMOTED_CATEGORIES:
            continue
        cd_ref, realm, rank, mode = resolve(row, by_cd_nom, by_cd_ref, by_name)
        resolution_modes[f"waiting:{mode}"] += 1
        if cd_ref is None or realm is None:
            continue
        replacement_refs.add(cd_ref)
        diagnostics["waitingDemotedRefs"] += 1

    candidates = diagnostics["matchedHarmonized"] + diagnostics["unmatched"] + diagnostics["ambiguous"]
    match_rate = diagnostics["matchedHarmonized"] / candidates if candidates else 1.0
    diagnostics["replacementRefs"] = len(replacement_refs)
    diagnostics["statuses"] = len(statuses)

    return {
        "schemaVersion": 1,
        "source": {
            "id": SOURCE_ID,
            "name": "Liste des espèces déterminantes ZNIEFF - Faune Grand Est",
            "producer": "DREAL Grand Est / CSRPN Grand Est / ODONAT Grand Est",
            "version": "LEDZfauna v2.2 - juin 2026",
            "publicationYear": 2026,
            "official": True,
            "checkedAt": checked_at,
            "sha256": digest,
            "landingPage": LANDING_URL,
            "sourceUrl": SOURCE_URL,
            "mirrorLandingPage": ODONAT_LANDING_URL,
            "canonicalSourceUrl": DREAL_URL,
        },
        "replaces": [{
            "region": "GES",
            "category": "znieff",
            "realm": "fauna",
            "cdRefs": sorted(replacement_refs),
        }],
        "statuses": sorted(statuses, key=lambda status: (
            status["cdRef"], status["label"], status.get("scopeLabel", ""), status["value"]
        )),
        "diagnostics": {
            **dict(diagnostics),
            "matchRate": round(match_rate, 6),
            "categories": dict(sorted(category_counts.items())),
            "waitingCategories": dict(sorted(waiting_counts.items())),
            "priorityValues": dict(sorted(priority_counts.items())),
            "resolutionModes": dict(sorted(resolution_modes.items())),
            "unresolvedSample": unresolved,
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--source", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.98)
    args = parser.parse_args()

    package = build_package(Path(args.taxref), Path(args.source), args.checked_at)
    diagnostics = package["diagnostics"]
    print(json.dumps(diagnostics, ensure_ascii=False, indent=2))
    if diagnostics["matchRate"] < args.min_match_rate:
        raise SystemExit(
            f"Taux de raccord TAXREF Grand Est insuffisant: {diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
        )
    if diagnostics["regionalDeterminantStatuses"] < 250:
        raise SystemExit(f"Volume EDZ Grand Est anormalement faible: {diagnostics['regionalDeterminantStatuses']}")
    if diagnostics["replacementRefs"] < 400:
        raise SystemExit(f"Couverture Grand Est anormalement faible: {diagnostics['replacementRefs']}")
    if not package["statuses"]:
        raise SystemExit("Aucun statut ZNIEFF Grand Est produit")

    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(f"Paquet régional écrit: {output} - {len(package['statuses'])} statuts")


if __name__ == "__main__":
    main()
