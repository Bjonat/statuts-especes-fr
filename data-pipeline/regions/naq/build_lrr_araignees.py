#!/usr/bin/env python3
"""Liste rouge Araignées Nouvelle-Aquitaine 2025 (évaluation UICN régionale)."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

SOURCE_ID = "dreal-naq-lrr-araignees-2025"
SOURCE_URL = (
    "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/tableau_final_evaluation.xlsx"
)
LANDING_URL = (
    "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/les-listes-rouges-regionales-a9991.html"
)
PRODUCER = "DREAL Nouvelle-Aquitaine / CSRPN / partenaires naturalistes"
EXPECTED_SHA256 = "abf2e14d8728c49626b49b38e5f8412659b076001f7befeec464960844f82fb3"
FILENAME = "araignees-2025.xlsx"
SHEET = "espèces évaluées"
VALID_LRR_CATEGORY = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[a-z]{0,3})$")
REALM_BY_KINGDOM = {"animalia": "fauna", "plantae": "flora"}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.replace("×", "x").casefold()


def core_name(value: object) -> str:
    text = clean(value).replace("×", "x")
    text = re.sub(r"\([^)]*\)", " ", text)
    tokens = [token for token in text.split() if not re.fullmatch(r"\d{4}", token) and token not in {",", "&"}]
    if len(tokens) >= 2:
        return normalize(f"{tokens[0]} {tokens[1]}")
    return normalize(text)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_category(value: object) -> str | None:
    text = clean(value)
    match = re.fullmatch(r"(?i)(EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA)([A-Za-z]{0,3})", text)
    if not match:
        return None
    base = match.group(1).upper()
    if base.startswith("CR") and "*" in match.group(1):
        base = "CR*"
    suffix = match.group(2)
    if base == "NA" and suffix:
        category = "NA" + suffix.lower()
    elif suffix:
        return None
    else:
        category = base
    return category if VALID_LRR_CATEGORY.fullmatch(category) else None


def read_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook[SHEET].iter_rows(values_only=True))
    workbook.close()
    headers = [clean(value) for value in rows[0]]
    if headers[:3] != ["Famille", "Nom scientifique", "Catégorie"]:
        raise RuntimeError(f"En-têtes inattendus: {headers[:3]}")
    parsed = []
    for values in rows[1:]:
        category = normalize_category(values[2] if len(values) > 2 else "")
        if category is None:
            continue
        name = clean(values[1] if len(values) > 1 else "")
        if not name or "/" in name:
            continue
        parsed.append({"name": name, "category": category})
    return parsed


def taxref_lookup(path: Path, wanted_names: set[str], wanted_cores: set[str]):
    by_name = defaultdict(set)
    by_core = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            if REALM_BY_KINGDOM.get(normalize(row.get("REGNE"))) != "fauna":
                continue
            cd_ref_raw = clean(row.get("CD_REF"))
            if not cd_ref_raw.isdigit():
                continue
            cd_ref = int(cd_ref_raw)
            for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                label = row.get(field)
                if not label:
                    continue
                key = normalize(label)
                if key in wanted_names:
                    by_name[key].add(cd_ref)
                core = core_name(label)
                if core in wanted_cores:
                    by_core[core].add(cd_ref)
    return by_name, by_core


def resolve(name: str, by_name, by_core):
    candidates = by_name.get(normalize(name), set())
    if not candidates:
        candidates = by_core.get(core_name(name), set())
    if len(candidates) == 1:
        return next(iter(candidates)), "name"
    if len(candidates) > 1:
        return None, "ambiguous"
    return None, "unmatched"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    path = Path(args.input_dir) / FILENAME
    digest = sha256(path)
    if digest != EXPECTED_SHA256:
        raise SystemExit(f"SHA-256 inattendu: {digest}")

    rows = read_rows(path)
    wanted_names = {normalize(row["name"]) for row in rows}
    wanted_cores = {core_name(row["name"]) for row in rows}
    by_name, by_core = taxref_lookup(Path(args.taxref), wanted_names, wanted_cores)

    stats = {"rows": len(rows), "matched": 0, "unmatched": 0, "ambiguous": 0, "unresolvedSample": [], "values": {}}
    values = defaultdict(int)
    statuses = []
    seen = set()
    for row in rows:
        cd_ref, mode = resolve(row["name"], by_name, by_core)
        if cd_ref is None:
            stats[mode] += 1
            if len(stats["unresolvedSample"]) < 30:
                stats["unresolvedSample"].append(
                    {"taxon": row["name"], "category": row["category"], "reason": mode}
                )
            continue
        stats["matched"] += 1
        values[row["category"]] += 1
        key = (cd_ref, row["category"])
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": cd_ref,
                "region": "NAQ",
                "category": "red_list_regional",
                "label": "Liste rouge régionale",
                "value": row["category"],
                "sourceId": SOURCE_ID,
                "scope": "regional",
            }
        )

    candidates = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    stats["matchRate"] = round(stats["matched"] / candidates, 6) if candidates else 1.0
    stats["values"] = dict(sorted(values.items()))
    if stats["matchRate"] < args.min_match_rate:
        raise SystemExit(f"Taux de raccord insuffisant: {stats['matchRate']:.2%}")
    if not statuses:
        raise SystemExit("Aucun statut araignées")

    covered = sorted({status["cdRef"] for status in statuses})
    package = {
        "schemaVersion": 1,
        "source": {
            "id": SOURCE_ID,
            "name": "Liste rouge Araignées Nouvelle-Aquitaine",
            "producer": PRODUCER,
            "version": "2025",
            "publicationYear": 2025,
            "official": True,
            "checkedAt": args.checked_at,
            "sha256": digest,
            "landingPage": LANDING_URL,
            "sourceUrl": SOURCE_URL,
        },
        "replaces": [
            {"region": "NAQ", "category": "red_list_regional", "realm": "fauna", "cdRefs": covered},
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": stats,
    }
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(f"Paquet écrit: {output} - {len(statuses)} statuts")


if __name__ == "__main__":
    main()
