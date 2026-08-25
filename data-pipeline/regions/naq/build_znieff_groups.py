#!/usr/bin/env python3
"""Adaptateurs ZNIEFF Nouvelle-Aquitaine — groupes hors flore vasculaire.

La flore vasculaire reste dans build_znieff.py (obv-na-znieff-flore-2019-v1.2).
Chaque groupe produit un paquet JSON distinct pour limiter le replaces BDC.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LANDING_URL = "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/les-listes-neo-aquitaines-a11234.html"
PRODUCER = "DREAL Nouvelle-Aquitaine / CSRPN Nouvelle-Aquitaine / partenaires naturalistes"
MAX_VALUE_LENGTH = 80
SEARCHABLE_RANKS = {"ES", "SSES", "VAR", "SVAR", "FO", "CAR", "RACE", "AGES"}
NAQ_DEPARTMENTS = {"16", "17", "19", "23", "24", "33", "40", "47", "64", "79", "86", "87"}
DEPT_LABELS = {
    "16": "Charente",
    "17": "Charente-Maritime",
    "19": "Corrèze",
    "23": "Creuse",
    "24": "Dordogne",
    "33": "Gironde",
    "40": "Landes",
    "47": "Lot-et-Garonne",
    "64": "Pyrénées-Atlantiques",
    "79": "Deux-Sèvres",
    "86": "Vienne",
    "87": "Haute-Vienne",
}
# Ordre des colonnes départementales oiseaux nicheurs (ligne d'en-tête 2).
OISEAUX_DEPT_COLUMNS = {
    9: "24",
    10: "33",
    11: "40",
    12: "47",
    13: "64",
    14: "19",
    15: "23",
    16: "87",
    17: "16",
    18: "17",
    19: "79",
    20: "86",
}

SOURCES = [
    {
        "key": "characees",
        "filename": "characees.xlsx",
        "id": "dreal-naq-znieff-characees-2023",
        "name": "ZNIEFF Characées Nouvelle-Aquitaine",
        "version": "2023",
        "year": 2023,
        "realm": "flora",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/cbn_2023-characees_determinantes_znieff_de_nouvelle-aquitaine.xlsx",
        "sha256": "0704cbff931ab70d42e7d738288884dd1ca098ec8d19787d2df013a4e26bcfc1",
    },
    {
        "key": "oiseaux-nicheurs",
        "filename": "oiseaux-nicheurs.xlsx",
        "id": "dreal-naq-znieff-oiseaux-nicheurs-2023",
        "name": "ZNIEFF Oiseaux nicheurs Nouvelle-Aquitaine",
        "version": "2023-03-16",
        "year": 2023,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/20230316_restitutiontableaued_oiseauxnich_synth.xlsx",
        "sha256": "93811416ef29b5c34d79a6e7d1cccdd32532a1f13408c238b6b125afe8460659",
    },
    {
        "key": "araignees",
        "filename": "araignees.xlsx",
        "id": "dreal-naq-znieff-araignees-2023",
        "name": "ZNIEFF Araignées Nouvelle-Aquitaine",
        "version": "2023",
        "year": 2023,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/listedeteraraigneesna.xlsx",
        "sha256": "5ad2961656e8567e3c0918ae9566895002f86713df5c7c22e8c507ae679a672d",
    },
    {
        "key": "amphibiens",
        "filename": "amphibiens.xlsx",
        "id": "dreal-naq-znieff-amphibiens-2024-09",
        "name": "ZNIEFF Amphibiens Nouvelle-Aquitaine",
        "version": "2024-09",
        "year": 2024,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/202409_listeespecesdeterminantesamphibiensnouvelle-aquitaine.xlsx",
        "sha256": "9f2e117eb0dd522c0a5efa64dc433d78a6086dd4e3cd16a363d0f62ea1f4b136",
    },
    {
        "key": "reptiles",
        "filename": "reptiles.xlsx",
        "id": "dreal-naq-znieff-reptiles-2024-09",
        "name": "ZNIEFF Reptiles Nouvelle-Aquitaine",
        "version": "2024-09",
        "year": 2024,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/202409_listeespecesdeterminantesreptilesnouvelle-aquitaine.xlsx",
        "sha256": "1f5511a53ba361a5a69ec79ab3788fce0be2254e427d64c9d6a9ac1a8f070529",
    },
    {
        "key": "mollusques",
        "filename": "mollusques.xlsx",
        "id": "dreal-naq-znieff-mollusques-2025",
        "name": "ZNIEFF Mollusques continentaux Nouvelle-Aquitaine",
        "version": "2025",
        "year": 2025,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/tab_mollusques.xlsx",
        "sha256": "ed3ea1b162406fe67ca1132a9fb1c0bdcb0a1e34e9019d96f0aa5d5c0ee17b94",
    },
    {
        "key": "orthopteres",
        "filename": "orthopteres.xlsx",
        "id": "dreal-naq-znieff-orthopteres-2026",
        "name": "ZNIEFF Orthoptères Nouvelle-Aquitaine",
        "version": "2026",
        "year": 2026,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/bonifait_duhaze-2026_orthopteres-determinants-znieff-na.xlsx",
        "sha256": "9fdcea3444ff414d2bc12b1654dd33a1d6d8041b3766091cc031445727323388",
    },
    {
        "key": "oiseaux-marins",
        "filename": "oiseaux-marins.xlsx",
        "id": "dreal-naq-znieff-oiseaux-marins-2026",
        "name": "ZNIEFF Oiseaux marins Nouvelle-Aquitaine",
        "version": "2026-06",
        "year": 2026,
        "realm": "fauna",
        "url": "https://www.nouvelle-aquitaine.developpement-durable.gouv.fr/IMG/xlsx/tableau_dz_oiseaux_marins.xlsx",
        "sha256": "aa139c515c210948e7f2acfe0a56f09f8782075cf70bdf3efbbf0aa65bcd5f64",
    },
]


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").replace("–", "-").replace("—", "-").replace("‑", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def mark(value: object) -> bool:
    return normalize(value) in {"x", "oui", "yes", "1", "true"}


def compact_value(value: object) -> str | None:
    text = clean(value)
    if not text or len(text) > MAX_VALUE_LENGTH:
        return None
    return text


def parse_departments(value: object) -> list[str]:
    text = clean(value).replace(";", ",").replace("/", ",").replace(" ", ",")
    # Typo amphibiens: "87.79" → 87 + 79
    text = text.replace(".", ",")
    codes: list[str] = []
    for token in re.split(r"[^0-9]+", text):
        if not token:
            continue
        code = token.zfill(2) if len(token) <= 2 else token
        if code in NAQ_DEPARTMENTS and code not in codes:
            codes.append(code)
    return codes


def scope_from_departments(departments: list[str]) -> tuple[str, str | None]:
    if not departments or set(departments) >= NAQ_DEPARTMENTS:
        return "regional", None
    labels = [DEPT_LABELS.get(code, code) for code in departments]
    return "partial", "Départements : " + ", ".join(labels)


def workbook_rows(path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"{path.name}: feuille absente {sheet_name}")
    return [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]


def add_candidate(candidates: list[dict], *, code: int | None, name: str, scope: str, scope_label: str | None = None, condition: str | None = None):
    if not code and not name:
        return
    candidates.append(
        {
            "code": code,
            "name": clean(name),
            "scope": scope,
            "scopeLabel": scope_label,
            "condition": compact_value(condition),
        }
    )


def parse_characees(path: Path) -> list[dict]:
    rows = workbook_rows(path, "Characées_ZNIEFF_N-A")
    header = [normalize(value) for value in rows[0]]
    cd_idx = header.index("cd_ref")
    name_idx = header.index("nom_valide")
    dept_idx = header.index("departement") if "departement" in header else None
    out: list[dict] = []
    for row in rows[1:]:
        code = as_int(row[cd_idx] if cd_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx < len(row) else "")
        if not code and not name:
            continue
        departments = parse_departments(row[dept_idx] if dept_idx is not None and dept_idx < len(row) else "")
        scope, scope_label = scope_from_departments(departments)
        add_candidate(out, code=code, name=name, scope=scope, scope_label=scope_label)
    return out


def parse_oiseaux_nicheurs(path: Path) -> list[dict]:
    rows = workbook_rows(path, "DéterminanceOiseauxNicheursPDF")
    out: list[dict] = []
    for row in rows[2:]:
        code = as_int(row[1] if len(row) > 1 else None)
        name = clean(row[6] if len(row) > 6 else "")
        regional = mark(row[7] if len(row) > 7 else None)
        departmental = mark(row[8] if len(row) > 8 else None)
        if not regional and not departmental:
            continue
        condition = clean(row[21] if len(row) > 21 else "")
        if regional:
            add_candidate(out, code=code, name=name, scope="regional", condition=condition)
            continue
        departments = [
            dept
            for index, dept in OISEAUX_DEPT_COLUMNS.items()
            if index < len(row) and mark(row[index])
        ]
        scope, scope_label = scope_from_departments(departments)
        add_candidate(out, code=code, name=name, scope=scope, scope_label=scope_label, condition=condition)
    return out


def parse_araignees(path: Path) -> list[dict]:
    rows = workbook_rows(path, "SP_ZNIEFF")
    header = [normalize(value) for value in rows[0]]
    code_idx = header.index("cd_nom")
    name_idx = header.index("nom_valide_taxref")
    statut_idx = header.index("statut znieff")
    condition_idx = header.index("condition") if "condition" in header else None
    out: list[dict] = []
    for row in rows[1:]:
        statut = normalize(row[statut_idx] if statut_idx < len(row) else "")
        if statut not in {"stricte", "sous conditions"}:
            continue
        code = as_int(row[code_idx] if code_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx < len(row) else "")
        condition = clean(row[condition_idx] if condition_idx is not None and condition_idx < len(row) else "")
        if condition:
            add_candidate(out, code=code, name=name, scope="partial", scope_label=condition, condition=condition)
        else:
            add_candidate(out, code=code, name=name, scope="regional")
    return out


def parse_herpeto(path: Path) -> list[dict]:
    rows = workbook_rows(path, "Feuil1")
    header = [normalize(value) for value in rows[0]]
    code_idx = next(i for i, value in enumerate(header) if value.startswith("cd ref"))
    ref_idx = header.index("espece de reference")
    cites_idx = header.index("nom(s) cite(s)") if "nom(s) cite(s)" in header else None
    dept_idx = header.index("departement")
    condition_idx = header.index("condition determinance") if "condition determinance" in header else None
    out: list[dict] = []
    for row in rows[1:]:
        code = as_int(row[code_idx] if code_idx < len(row) else None)
        ref_name = clean(row[ref_idx] if ref_idx < len(row) else "")
        cited = clean(row[cites_idx] if cites_idx is not None and cites_idx < len(row) else "")
        # Amphibiens: CD_REF + vernacular in Espèce + scientific in Nom(s) cité(s)
        # Reptiles: CD_REF + scientific in Espèce + vernacular in Nom(s) cité(s)
        scientific = cited if (cited and re.search(r"[a-z]\s+[a-z]", normalize(cited))) else ref_name
        if not code and not scientific:
            continue
        departments = parse_departments(row[dept_idx] if dept_idx < len(row) else "")
        scope, scope_label = scope_from_departments(departments)
        condition = clean(row[condition_idx] if condition_idx is not None and condition_idx < len(row) else "")
        add_candidate(out, code=code, name=scientific, scope=scope, scope_label=scope_label, condition=condition)
    return out


def parse_mollusques(path: Path) -> list[dict]:
    rows = workbook_rows(path, "LISTE")
    header = [normalize(value) for value in rows[0]]
    code_idx = header.index("cd_ref_taxref")
    name_idx = header.index("nom_valide_taxref")
    det_idx = header.index("det_znieff")
    out: list[dict] = []
    for row in rows[1:]:
        det = clean(row[det_idx] if det_idx < len(row) else "").upper()
        if det != "ZNIEFF":
            continue
        code = as_int(row[code_idx] if code_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx < len(row) else "")
        add_candidate(out, code=code, name=name, scope="regional")
    return out


def parse_orthopteres(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    regional_rows = [list(row) for row in workbook["Liste-régionale"].iter_rows(values_only=True)]
    exception_rows = [list(row) for row in workbook["Exceptions-départementales"].iter_rows(values_only=True)]
    out: list[dict] = []
    seen_codes: set[int] = set()

    for row in regional_rows[3:]:
        code = as_int(row[0] if row else None)
        name = clean(row[1] if len(row) > 1 else "")
        determinant = normalize(row[18] if len(row) > 18 else "")
        if determinant != "oui":
            continue
        condition = clean(row[21] if len(row) > 21 else "")
        add_candidate(out, code=code, name=name, scope="regional", condition=condition)
        if code:
            seen_codes.add(code)

    for row in exception_rows[1:]:
        code = as_int(row[0] if row else None)
        name = clean(row[1] if len(row) > 1 else "")
        depts_raw = clean(row[2] if len(row) > 2 else "")
        if not code and not name:
            continue
        # "Deux-Sèvres, Vienne" → 79, 86
        departments: list[str] = []
        lowered = normalize(depts_raw)
        for code_dept, label in DEPT_LABELS.items():
            if normalize(label) in lowered:
                departments.append(code_dept)
        departments = parse_departments(depts_raw) or departments
        scope, scope_label = scope_from_departments(departments)
        condition = clean(row[3] if len(row) > 3 else "")
        if code in seen_codes:
            continue
        add_candidate(out, code=code, name=name, scope=scope, scope_label=scope_label, condition=condition)
    return out


def parse_oiseaux_marins(path: Path) -> list[dict]:
    rows = workbook_rows(path, "Espèces déterminantes")
    out: list[dict] = []
    for row in rows[2:]:
        name = clean(row[0] if row else "")
        if not name or normalize(name) in {"nom latin"}:
            continue
        add_candidate(out, code=None, name=name, scope="regional")
    return out


PARSERS = {
    "characees": parse_characees,
    "oiseaux-nicheurs": parse_oiseaux_nicheurs,
    "araignees": parse_araignees,
    "amphibiens": parse_herpeto,
    "reptiles": parse_herpeto,
    "mollusques": parse_mollusques,
    "orthopteres": parse_orthopteres,
    "oiseaux-marins": parse_oiseaux_marins,
}


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str], realms: set[str]):
    by_cd_nom: dict[int, tuple[int, str, str]] = {}
    by_cd_ref: dict[int, tuple[int, str, str]] = {}
    by_name: dict[str, set[tuple[int, str, str]]] = defaultdict(set)

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom = as_int(row.get("CD_NOM"))
            cd_ref = as_int(row.get("CD_REF"))
            if not cd_nom or not cd_ref:
                continue
            kingdom = normalize(row.get("REGNE"))
            realm = {"animalia": "fauna", "plantae": "flora"}.get(kingdom)
            if realm not in realms:
                continue
            rank = clean(row.get("RANG")).upper()
            entry = (cd_ref, realm, rank)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = entry
            if cd_ref in wanted_codes and cd_nom == cd_ref:
                by_cd_ref[cd_ref] = entry
            for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                label = clean(row.get(field))
                key = normalize(label)
                if label and key in wanted_names:
                    by_name[key].add(entry)
    return by_cd_nom, by_cd_ref, by_name


def resolve(candidate: dict, by_cd_nom, by_cd_ref, by_name, expected_realm: str):
    code = candidate["code"]
    if code and code in by_cd_nom:
        cd_ref, realm, rank = by_cd_nom[code]
        if realm != expected_realm:
            return None, None, "excluded_realm"
        if rank and rank not in SEARCHABLE_RANKS:
            return None, None, "excluded_rank"
        return cd_ref, realm, "cd_nom"
    if code and code in by_cd_ref:
        cd_ref, realm, rank = by_cd_ref[code]
        if realm != expected_realm:
            return None, None, "excluded_realm"
        if rank and rank not in SEARCHABLE_RANKS:
            return None, None, "excluded_rank"
        return cd_ref, realm, "cd_ref"

    name = candidate["name"]
    if name:
        matches = {
            entry
            for entry in by_name.get(normalize(name), set())
            if entry[1] == expected_realm and (not entry[2] or entry[2] in SEARCHABLE_RANKS)
        }
        if len(matches) == 1:
            cd_ref, realm, _rank = next(iter(matches))
            return cd_ref, realm, "name"
        if len(matches) > 1:
            return None, None, "ambiguous"
    return None, None, "unmatched"


def emit_statuses(source: dict, candidates: list[dict], by_cd_nom, by_cd_ref, by_name):
    stats = Counter()
    unresolved = []
    statuses = []
    seen = set()
    expected_realm = source["realm"]

    for candidate in candidates:
        stats["rows"] += 1
        cd_ref, realm, mode = resolve(candidate, by_cd_nom, by_cd_ref, by_name, expected_realm)
        if cd_ref is None:
            stats[mode] += 1
            if len(unresolved) < 40:
                unresolved.append(
                    {"code": candidate["code"], "taxon": candidate["name"], "reason": mode}
                )
            continue
        stats["matched"] += 1
        stats[mode] += 1

        records = [
            {
                "cdRef": cd_ref,
                "region": "NAQ",
                "category": "znieff",
                "label": "Déterminante ZNIEFF",
                "value": "Oui",
                "sourceId": source["id"],
                "scope": candidate["scope"],
            }
        ]
        if candidate.get("scopeLabel"):
            records[0]["scopeLabel"] = candidate["scopeLabel"]
        if candidate.get("condition"):
            records.append(
                {
                    "cdRef": cd_ref,
                    "region": "NAQ",
                    "category": "znieff",
                    "label": "Condition de déterminance",
                    "value": candidate["condition"],
                    "sourceId": source["id"],
                    "scope": candidate["scope"],
                    **({"scopeLabel": candidate["scopeLabel"]} if candidate.get("scopeLabel") else {}),
                }
            )

        for record in records:
            key = (
                record["cdRef"],
                record["label"],
                record["value"],
                record["scope"],
                record.get("scopeLabel", ""),
            )
            if key in seen:
                continue
            seen.add(key)
            statuses.append(record)

    candidates_n = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    match_rate = round(stats["matched"] / candidates_n, 6) if candidates_n else 1.0
    return statuses, {
        **{key: int(value) for key, value in stats.items()},
        "matchRate": match_rate,
        "unresolvedSample": unresolved,
    }


def build_package(source: dict, input_dir: Path, by_cd_nom, by_cd_ref, by_name, checked_at: str):
    path = input_dir / source["filename"]
    digest = sha256(path)
    if digest != source["sha256"]:
        raise RuntimeError(f"{source['id']}: SHA-256 inattendu {digest}")

    candidates = PARSERS[source["key"]](path)
    statuses, diagnostics = emit_statuses(source, candidates, by_cd_nom, by_cd_ref, by_name)
    covered = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": digest,
            "landingPage": LANDING_URL,
            "sourceUrl": source["url"],
        },
        "replaces": [
            {
                "region": "NAQ",
                "category": "znieff",
                "realm": source["realm"],
                "cdRefs": covered,
            }
        ],
        "statuses": sorted(
            statuses,
            key=lambda status: (status["cdRef"], status["label"], status["value"], status.get("scopeLabel", "")),
        ),
        "diagnostics": diagnostics,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    parser.add_argument(
        "--only",
        nargs="*",
        default=None,
        help="Limiter aux clés de groupes (characees, amphibiens, …)",
    )
    args = parser.parse_args()

    sources = [source for source in SOURCES if args.only is None or source["key"] in args.only]
    if not sources:
        raise SystemExit("Aucun groupe sélectionné")

    parsed: dict[str, list[dict]] = {}
    wanted_codes: set[int] = set()
    wanted_names: set[str] = set()
    realms: set[str] = set()
    for source in sources:
        path = Path(args.input_dir) / source["filename"]
        if not path.exists():
            raise SystemExit(f"Fichier manquant: {path}")
        candidates = PARSERS[source["key"]](path)
        parsed[source["key"]] = candidates
        realms.add(source["realm"])
        for candidate in candidates:
            if candidate["code"]:
                wanted_codes.add(candidate["code"])
            if candidate["name"]:
                wanted_names.add(normalize(candidate["name"]))

    by_cd_nom, by_cd_ref, by_name = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names, realms)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    total = 0
    for source in sources:
        package = build_package(
            source,
            Path(args.input_dir),
            by_cd_nom,
            by_cd_ref,
            by_name,
            args.checked_at,
        )
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        if "url" in package["source"]:
            raise SystemExit(f"{source['id']}: champ url interdit dans la source publiée")
        long_values = [status for status in package["statuses"] if len(status["value"]) > MAX_VALUE_LENGTH]
        if long_values:
            raise SystemExit(f"{source['id']}: {len(long_values)} valeurs > {MAX_VALUE_LENGTH} caractères")

        output = out_dir / f"naq-znieff-{source['key']}.json"
        output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        total += len(package["statuses"])
        print(f"Paquet écrit: {output} — {len(package['statuses'])} statuts")

    print(f"NAQ ZNIEFF groupes: {len(sources)} paquets, {total} statuts")


if __name__ == "__main__":
    main()
