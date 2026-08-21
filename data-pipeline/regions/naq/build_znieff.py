#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

RANK_MARKERS = {
    'subsp.', 'subsp', 'ssp.', 'ssp', 'var.', 'var', 'f.', 'f', 'fo.', 'fo',
    'subvar.', 'subvar', 'nothosubsp.', 'nothosubsp',
}

DEPARTMENT_COLUMNS = {
    3: '24', 4: '47', 5: '33', 6: '40', 7: '64', 8: '16',
    9: '17', 10: '79', 11: '86', 12: '19', 13: '87', 14: '23',
}


def normalize(value: object) -> str:
    text = unicodedata.normalize('NFKD', str(value or ''))
    text = ''.join(char for char in text if not unicodedata.combining(char))
    text = text.replace('×', 'x')
    text = re.sub(r'\s+', ' ', text).strip().casefold()
    return text


def core_name(value: object) -> str:
    tokens = str(value or '').replace('×', 'x').strip().split()
    if len(tokens) < 2:
        return normalize(value)
    keep = tokens[:2]
    if len(tokens) >= 4 and tokens[2].casefold() in RANK_MARKERS:
        keep.extend(tokens[2:4])
    return normalize(' '.join(keep))


def header_value(row: dict[str, str], *names: str) -> str:
    lowered = {str(key).strip().casefold(): value for key, value in row.items()}
    for name in names:
        if name.casefold() in lowered:
            return str(lowered[name.casefold()] or '').strip()
    return ''


def taxref_lookup(path: Path):
    exact = defaultdict(set)
    bare = defaultdict(set)
    with path.open('r', encoding='utf-8-sig', newline='') as handle:
        reader = csv.DictReader(handle, delimiter='\t')
        for row in reader:
            kingdom = header_value(row, 'REGNE')
            if kingdom.casefold() != 'plantae':
                continue
            cd_ref_raw = header_value(row, 'CD_REF')
            if not cd_ref_raw.isdigit():
                continue
            cd_ref = int(cd_ref_raw)
            for field in ('LB_NOM', 'NOM_COMPLET', 'NOM_VALIDE'):
                label = header_value(row, field)
                if label:
                    exact[normalize(label)].add(cd_ref)
            lb_nom = header_value(row, 'LB_NOM')
            if lb_nom:
                bare[normalize(lb_nom)].add(cd_ref)
    return exact, bare


def resolve_taxon(name: str, exact, bare):
    candidates = exact.get(normalize(name), set())
    if len(candidates) == 1:
        return next(iter(candidates)), 'exact'
    candidates = bare.get(core_name(name), set())
    if len(candidates) == 1:
        return next(iter(candidates)), 'core'
    if len(candidates) > 1:
        return None, 'ambiguous'
    return None, 'unmatched'


def x_mark(value: object) -> bool:
    return normalize(value) == 'x'


def build_package(taxref_path: Path, workbook_path: Path, checked_at: str):
    exact, bare = taxref_lookup(taxref_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if 'Liste Determinantes ZNIEFF' not in workbook.sheetnames:
        raise RuntimeError('Onglet "Liste Determinantes ZNIEFF" introuvable')
    sheet = workbook['Liste Determinantes ZNIEFF']

    statuses = []
    seen = set()
    stats = {'rows': 0, 'determining': 0, 'exact': 0, 'core': 0, 'unmatched': 0, 'ambiguous': 0}
    unresolved = []

    for values in sheet.iter_rows(min_row=3, values_only=True):
        taxon_name = str(values[0] or '').strip()
        if not taxon_name:
            continue
        stats['rows'] += 1
        regional = x_mark(values[1])
        departmental = x_mark(values[2])
        if not regional and not departmental:
            continue
        stats['determining'] += 1

        cd_ref, mode = resolve_taxon(taxon_name, exact, bare)
        stats[mode] += 1
        if cd_ref is None:
            unresolved.append({'taxon': taxon_name, 'reason': mode})
            continue

        restriction = str(values[22] or '').strip() if len(values) > 22 else ''
        departments = [code for index, code in DEPARTMENT_COLUMNS.items() if index < len(values) and x_mark(values[index])]

        scope = 'regional'
        scope_label = None
        if regional and restriction:
            scope = 'partial'
            scope_label = restriction
        elif not regional and departmental:
            scope = 'partial'
            scope_label = f"Départements : {', '.join(departments)}" if departments else 'Déterminance départementale'

        key = (cd_ref, scope, scope_label)
        if key in seen:
            continue
        seen.add(key)
        status = {
            'cdRef': cd_ref,
            'region': 'NAQ',
            'category': 'znieff',
            'label': 'Déterminante ZNIEFF',
            'value': 'Oui',
            'sourceId': 'obv-na-znieff-flore-2019-v1.2',
            'scope': scope,
        }
        if scope_label:
            status['scopeLabel'] = scope_label
        statuses.append(status)

    matched = stats['exact'] + stats['core']
    match_rate = matched / stats['determining'] if stats['determining'] else 0
    covered_refs = sorted({status['cdRef'] for status in statuses})
    package = {
        'schemaVersion': 1,
        'source': {
            'id': 'obv-na-znieff-flore-2019-v1.2',
            'name': 'ZNIEFF flore Nouvelle-Aquitaine',
            'producer': 'CBN de Nouvelle-Aquitaine / CSRPN Nouvelle-Aquitaine',
            'version': 'v1.2 - 2019',
            'publicationYear': 2019,
            'official': True,
            'checkedAt': checked_at,
        },
        'replaces': [
            {'region': 'NAQ', 'category': 'znieff', 'realm': 'flora', 'cdRefs': covered_refs},
        ],
        'statuses': sorted(statuses, key=lambda status: (status['cdRef'], status.get('scopeLabel', ''))),
        'diagnostics': {
            **stats,
            'matched': matched,
            'matchRate': round(match_rate, 6),
            'unresolvedSample': unresolved[:50],
        },
    }
    return package


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--taxref', required=True)
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--out', required=True)
    parser.add_argument('--checked-at', default=date.today().isoformat())
    parser.add_argument('--min-match-rate', type=float, default=0.97)
    args = parser.parse_args()

    package = build_package(Path(args.taxref), Path(args.workbook), args.checked_at)
    diagnostics = package['diagnostics']
    print(json.dumps(diagnostics, ensure_ascii=False, indent=2))
    if diagnostics['matchRate'] < args.min_match_rate:
        raise SystemExit(f"Taux de correspondance TAXREF insuffisant: {diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}")

    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(package, ensure_ascii=False, separators=(',', ':')) + '\n', encoding='utf-8')
    print(f"Paquet régional écrit: {output} - {len(package['statuses'])} statuts")


if __name__ == '__main__':
    main()