#!/usr/bin/env python3
"""ZNIEFF Provence-Alpes-Côte d'Azur — faune 2024 + flore 2016."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import xlrd
from openpyxl import load_workbook

LANDING_URL = "https://www.paca.developpement-durable.gouv.fr/les-listes-rouges-regionales-a7296.html"
ZNIEFF_LANDING = "https://www.paca.developpement-durable.gouv.fr/"
PRODUCER = "DREAL Provence-Alpes-Côte d'Azur / CSRPN PACA / partenaires naturalistes"
MAX_VALUE_LENGTH = 80
SEARCHABLE_RANKS = {"ES", "SSES", "VAR", "SVAR", "FO", "CAR", "RACE", "AGES"}

SOURCES = [
    {
        "key": "fauna",
        "filename": "znieff-fauna-2024.xlsx",
        "id": "dreal-pac-znieff-fauna-2024-01",
        "name": "ZNIEFF Faune Provence-Alpes-Côte d'Azur",
        "version": "2024-01",
        "year": 2024,
        "realm": "fauna",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/znieff_faune_janv-2024.xlsx",
        "sha256": "d38ffb58944a998ac937146fcdeed606b0e40cd178cdc2bf467726f81a672375",
    },
    {
        "key": "flora",
        "filename": "znieff-flora-2016.xls",
        "id": "dreal-pac-znieff-flora-2016",
        "name": "ZNIEFF Flore Provence-Alpes-Côte d'Azur",
        "version": "2016",
        "year": 2016,
        "realm": "flora",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xls/znieff_flore_2016.xls",
        "sha256": "1c39c39f36ca97659e6b534d7e5a7c385ea3e9e470679c6d360f6e3cc8d6204c",
    },
]


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").replace("–", "-").replace("—", "-").replace("‑", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if text.startswith("#"):
        return None
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def znieff_value(raw: object) -> str | None:
    key = normalize(raw)
    if key in {"determinante", "determinant"}:
        return "Déterminante"
    if key == "remarquable":
        return "Remarquable"
    return None


def parse_fauna(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["faune"].iter_rows(values_only=True))
    header = [normalize(value) for value in rows[0]]
    cd_ref_idx = header.index("cd_ref_taxref_v16")
    cd_nom_idx = header.index("cd_nom_taxref_v16")
    name_idx = header.index("nom_valide_taxref_v16")
    statut_idx = header.index("statut znieff")
    rang_idx = header.index("rang")
    out: list[dict] = []
    for row in rows[1:]:
        if not any(row):
            continue
        value = znieff_value(row[statut_idx] if statut_idx < len(row) else None)
        if not value:
            continue
        code = as_int(row[cd_ref_idx] if cd_ref_idx < len(row) else None)
        if code is None:
            code = as_int(row[cd_nom_idx] if cd_nom_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx < len(row) else "")
        if name.startswith("#") or name.upper() == "N/A":
            name = ""
        rank = clean(row[rang_idx] if rang_idx < len(row) else "").upper()
        if not code and not name:
            continue
        out.append({"code": code, "name": name, "value": value, "rank": rank})
    return out


def parse_flora(path: Path) -> list[dict]:
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_name("Liste_sp_ZNIEFF_Flore")
    header = [normalize(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
    code_idx = header.index("cd_nom_tax ref 5")
    name_idx = header.index("libelle_cd_nom_tax ref 5")
    statut_idx = next(index for index, value in enumerate(header) if "znieff" in value)
    rang_idx = header.index("rang")
    best: dict[int | str, dict] = {}
    priority = {"Déterminante": 2, "Remarquable": 1}
    for row_index in range(1, sheet.nrows):
        value = znieff_value(sheet.cell_value(row_index, statut_idx))
        if not value:
            continue
        code = as_int(sheet.cell_value(row_index, code_idx))
        name = clean(sheet.cell_value(row_index, name_idx))
        rank = clean(sheet.cell_value(row_index, rang_idx)).upper()
        if not code and not name:
            continue
        key = code if code is not None else normalize(name)
        candidate = {"code": code, "name": name, "value": value, "rank": rank}
        previous = best.get(key)
        if previous is None or priority[value] > priority[previous["value"]]:
            best[key] = candidate
    return list(best.values())


PARSERS = {"fauna": parse_fauna, "flora": parse_flora}


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str], realms: set[str]):
    by_cd_nom: dict[int, tuple[int, str, str]] = {}
    by_cd_ref: dict[int, tuple[int, str, str]] = {}
    by_name: dict[str, set[tuple[int, str, str]]] = defaultdict(set)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom = as_int(row.get("CD_NOM"))
            cd_ref = as_int(row.get("CD_REF"))
            if not cd_nom or not cd_ref:
                continue
            realm = {"animalia": "fauna", "plantae": "flora"}.get(normalize(row.get("REGNE")))
            if realm not in realms:
                continue
            rank = clean(row.get("RANG")).upper()
            entry = (cd_ref, realm, rank)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = entry
            if cd_ref in wanted_codes and cd_nom == cd_ref:
                by_cd_ref[cd_ref] = entry
            for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                label = clean(row.get(field))
                key = normalize(label)
                if label and key in wanted_names:
                    by_name[key].add(entry)
    return by_cd_nom, by_cd_ref, by_name


def resolve(candidate, by_cd_nom, by_cd_ref, by_name, expected_realm: str):
    code = candidate["code"]
    if code and code in by_cd_nom:
        cd_ref, realm, rank = by_cd_nom[code]
        if realm != expected_realm:
            return None, "excluded_realm"
        if rank and rank not in SEARCHABLE_RANKS:
            return None, "excluded_rank"
        return cd_ref, "cd_nom"
    if code and code in by_cd_ref:
        cd_ref, realm, rank = by_cd_ref[code]
        if realm != expected_realm:
            return None, "excluded_realm"
        if rank and rank not in SEARCHABLE_RANKS:
            return None, "excluded_rank"
        return cd_ref, "cd_ref"
    name = candidate["name"]
    if name:
        matches = {
            entry
            for entry in by_name.get(normalize(name), set())
            if entry[1] == expected_realm and (not entry[2] or entry[2] in SEARCHABLE_RANKS)
        }
        if len(matches) == 1:
            return next(iter(matches))[0], "name"
        if len(matches) > 1:
            return None, "ambiguous"
    return None, "unmatched"


def build_package(source, input_dir: Path, by_cd_nom, by_cd_ref, by_name, checked_at: str):
    path = input_dir / source["filename"]
    digest = sha256(path)
    if digest != source["sha256"]:
        raise RuntimeError(f"{source['id']}: SHA-256 inattendu {digest}")
    candidates = PARSERS[source["key"]](path)
    stats = Counter()
    unresolved = []
    statuses = []
    seen = set()
    for candidate in candidates:
        stats["rows"] += 1
        cd_ref, mode = resolve(candidate, by_cd_nom, by_cd_ref, by_name, source["realm"])
        if cd_ref is None:
            stats[mode] += 1
            if len(unresolved) < 40:
                unresolved.append({"code": candidate["code"], "taxon": candidate["name"], "reason": mode})
            continue
        stats["matched"] += 1
        stats[mode] += 1
        value = candidate["value"]
        if len(value) > MAX_VALUE_LENGTH:
            stats["omittedLong"] += 1
            continue
        key = (cd_ref, value)
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": cd_ref,
                "region": "PAC",
                "category": "znieff",
                "label": "Statut ZNIEFF",
                "value": value,
                "sourceId": source["id"],
                "scope": "regional",
            }
        )
    candidates_n = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    match_rate = round(stats["matched"] / candidates_n, 6) if candidates_n else 1.0
    covered = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": digest,
            "landingPage": ZNIEFF_LANDING,
            "sourceUrl": source["url"],
        },
        "replaces": [{"region": "PAC", "category": "znieff", "realm": source["realm"], "cdRefs": covered}],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": {
            **{key: int(value) for key, value in stats.items()},
            "matchRate": match_rate,
            "unresolvedSample": unresolved,
            "values": dict(Counter(status["value"] for status in statuses)),
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    wanted_codes: set[int] = set()
    wanted_names: set[str] = set()
    realms: set[str] = set()
    parsed: dict[str, list[dict]] = {}
    for source in SOURCES:
        path = Path(args.input_dir) / source["filename"]
        candidates = PARSERS[source["key"]](path)
        parsed[source["key"]] = candidates
        realms.add(source["realm"])
        for candidate in candidates:
            if candidate["code"]:
                wanted_codes.add(candidate["code"])
            if candidate["name"]:
                wanted_names.add(normalize(candidate["name"]))

    by_cd_nom, by_cd_ref, by_name = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names, realms)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    total = 0
    for source in SOURCES:
        package = build_package(source, Path(args.input_dir), by_cd_nom, by_cd_ref, by_name, args.checked_at)
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        if "url" in package["source"]:
            raise SystemExit(f"{source['id']}: champ url interdit")
        output = out_dir / f"pac-znieff-{source['key']}.json"
        output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        total += len(package["statuses"])
        print(f"Paquet écrit: {output} — {len(package['statuses'])} statuts")
    print(f"PACA ZNIEFF: {len(SOURCES)} paquets, {total} statuts")


if __name__ == "__main__":
    main()
