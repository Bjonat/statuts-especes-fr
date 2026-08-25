#!/usr/bin/env python3
"""Listes rouges régionales Provence-Alpes-Côte d'Azur — multi-groupes."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LANDING_URL = "https://www.paca.developpement-durable.gouv.fr/listes-rouges-regionales-a7296.html"
PRODUCER = "DREAL Provence-Alpes-Côte d'Azur / CSRPN PACA / partenaires naturalistes"
MAX_VALUE_LENGTH = 80
SEARCHABLE_RANKS = {"ES", "SSES", "VAR", "SVAR", "FO", "CAR", "RACE", "AGES"}
VALID_LRR = re.compile(r"^(?:EX|EW|RE|CR\*?|EN|VU|NT|LC|DD|NE|NA[a-dA-D]?)$")

SOURCES = [
    {
        "key": "oiseaux",
        "filename": "lrr-oiseaux-2020.xlsx",
        "id": "dreal-pac-lrr-oiseaux-2020",
        "name": "Liste rouge Oiseaux Provence-Alpes-Côte d'Azur",
        "version": "2020",
        "year": 2020,
        "realm": "fauna",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/liste_rouge_avifaune_paca_csrpn_janv2020.xlsx",
        "sha256": "c63509fe1ac9a0c444c6771bc5913d4f0749a8adca7f52985859faf155ea8369",
    },
    {
        "key": "odonates",
        "filename": "lrr-odonates-2017.xlsx",
        "id": "dreal-pac-lrr-odonates-2017",
        "name": "Liste rouge Odonates Provence-Alpes-Côte d'Azur",
        "version": "2017",
        "year": 2017,
        "realm": "fauna",
        "taxrefOrder": "Odonata",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/lr_paca_odonates_2017_web.xlsx",
        "sha256": "bf80fe46a613e6a1002802b8e919526669ee0bce01e69a901d8264411bf00fc6",
    },
    {
        "key": "papillons",
        "filename": "lrr-papillons-2024.xlsx",
        "id": "dreal-pac-lrr-papillons-2024",
        "name": "Liste rouge Papillons Provence-Alpes-Côte d'Azur",
        "version": "2024",
        "year": 2024,
        "realm": "fauna",
        "taxrefOrder": "Lepidoptera",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/tableau_simplifie_lrr-pap_2024.xlsx",
        "sha256": "937ac5296de2dac5118013c63b22fe5cc4564e2520a3d4506a0ece0d2142e973",
    },
    {
        "key": "flore",
        "filename": "lrr-flore-2015.xlsx",
        "id": "dreal-pac-lrr-flore-2015",
        "name": "Liste rouge Flore vasculaire Provence-Alpes-Côte d'Azur",
        "version": "2015",
        "year": 2015,
        "realm": "flora",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/LR_PACA_Flore_2015_web.xlsx",
        "sha256": "99b44c10f572521e83999543abf14a9c38c75d9a56a4c2540c1d6dc72c9a440a",
    },
    {
        "key": "amphibiens",
        "filename": "lrr-herpeto-2016.xlsx",
        "id": "dreal-pac-lrr-amphibiens-2016",
        "name": "Liste rouge Amphibiens Provence-Alpes-Côte d'Azur",
        "version": "2016",
        "year": 2016,
        "realm": "fauna",
        "sheet": "amphibiens",
        "taxrefClass": "Amphibia",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/lrr_paca_reptiles_amphibiens_2016.xlsx",
        "sha256": "fe81c5f6a4e3708838e8bb85e15df389e02970e6ca63c86201a9b9d556790ba9",
    },
    {
        "key": "reptiles",
        "filename": "lrr-herpeto-2016.xlsx",
        "id": "dreal-pac-lrr-reptiles-2016",
        "name": "Liste rouge Reptiles Provence-Alpes-Côte d'Azur",
        "version": "2016",
        "year": 2016,
        "realm": "fauna",
        "sheet": "reptiles",
        "taxrefClass": "Reptilia",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/lrr_paca_reptiles_amphibiens_2016.xlsx",
        "sha256": "fe81c5f6a4e3708838e8bb85e15df389e02970e6ca63c86201a9b9d556790ba9",
    },
    {
        "key": "orthopteres",
        "filename": "lrr-orthopteres-2018.xlsx",
        "id": "dreal-pac-lrr-orthopteres-2018",
        "name": "Liste rouge Orthoptères Provence-Alpes-Côte d'Azur",
        "version": "2018",
        "year": 2018,
        "realm": "fauna",
        "taxrefOrder": "Orthoptera",
        "url": "https://www.paca.developpement-durable.gouv.fr/IMG/xlsx/lr_paca_orthop_2018_web.xlsx",
        "sha256": "5d07840fadce9dab531f9317ead43ad0b7f3afb1e2c702ae48583d32687364ad",
    },
]


def clean(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").replace("–", "-").replace("—", "-").replace("‑", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("×", "x")
    return text.casefold()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return int(float(text))
    return None


def normalize_category(value: object) -> str | None:
    text = clean(value).replace(" ", "")
    if not text:
        return None
    upper = text.upper()
    na_map = {"NAA": "NAa", "NAB": "NAb", "NAC": "NAc", "NAD": "NAd", "NA": "NA"}
    if upper in na_map:
        return na_map[upper]
    if upper.startswith("CR") and "*" in text:
        return "CR*"
    if re.fullmatch(r"(?:EX|EW|RE|CR|EN|VU|NT|LC|DD|NE)", upper):
        return upper
    return None


def strip_authorship(name: str) -> str:
    text = clean(name)
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text)
    text = re.sub(r",\s*\d{4}\s*$", "", text)
    return clean(text)


def workbook_rows(path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"{path.name}: feuille absente {sheet_name}")
    return [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]


def find_header_row(rows: list[list[object]], required_substrings: list[str]) -> int:
    required = [normalize(value) for value in required_substrings]
    for index, row in enumerate(rows[:20]):
        headers = [normalize(value) for value in row]
        if all(any(target in header for header in headers if header) for target in required):
            return index
    raise RuntimeError(f"En-tête introuvable ({required_substrings})")


def column_index(headers: list[str], *candidates: str) -> int | None:
    for candidate in candidates:
        target = normalize(candidate)
        for index, header in enumerate(headers):
            if header == target or target in header:
                return index
    return None


def parse_oiseaux(path: Path) -> list[dict]:
    rows = workbook_rows(path, "Referentiel LRR")
    header = [normalize(value) for value in rows[0]]
    code_idx = column_index(header, "cd_ref")
    name_idx = column_index(header, "nom scientifique", "nom_valide")
    cat_idx = next(index for index, value in enumerate(header) if "liste rouge paca (2020)" in value)
    out = []
    for row in rows[1:]:
        category = normalize_category(row[cat_idx] if cat_idx < len(row) else None)
        if not category:
            continue
        out.append(
            {
                "code": as_int(row[code_idx] if code_idx is not None and code_idx < len(row) else None),
                "name": clean(row[name_idx] if name_idx is not None and name_idx < len(row) else ""),
                "category": category,
            }
        )
    return out


def parse_odonates(path: Path) -> list[dict]:
    rows = workbook_rows(path, "LRR-Paca-Odonates 2017")
    header_row = find_header_row(rows, ["taxon", "region paca (2017)"])
    headers = [normalize(value) for value in rows[header_row]]
    name_idx = column_index(headers, "taxon")
    cat_idx = column_index(headers, "region paca (2017)")
    out = []
    for row in rows[header_row + 1 :]:
        category = normalize_category(row[cat_idx] if cat_idx is not None and cat_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        if not category or not name:
            continue
        out.append({"code": None, "name": name, "category": category})
    return out


def parse_papillons(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet_name in ("esp", "ssp"):
        rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
        headers = [normalize(value) for value in rows[0]]
        code_idx = column_index(headers, "cd_ref")
        name_idx = column_index(headers, "taxon", "nom scientifique")
        cat_idx = next(index for index, value in enumerate(headers) if "lr paca reevaluation 2024" in value or "lr paca réévaluation 2024" in value)
        for row in rows[1:]:
            if not any(row):
                continue
            category = normalize_category(row[cat_idx] if cat_idx < len(row) else None)
            if not category:
                continue
            out.append(
                {
                    "code": as_int(row[code_idx] if code_idx is not None and code_idx < len(row) else None),
                    "name": clean(row[name_idx] if name_idx is not None and name_idx < len(row) else ""),
                    "category": category,
                }
            )
    return out


def parse_flore(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    out = []
    # especes_com
    rows = [list(row) for row in workbook["especes_com"].iter_rows(values_only=True)]
    header_row = find_header_row(rows, ["espece", "lr"])
    headers = [normalize(value) for value in rows[header_row]]
    name_idx = column_index(headers, "espece", "taxon")
    cat_idx = next(index for index, value in enumerate(headers) if value.startswith("lr"))
    for row in rows[header_row + 1 :]:
        category = normalize_category(row[cat_idx] if cat_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        if not category or not name:
            continue
        out.append({"code": None, "name": name, "category": category})
    # infra_com
    rows = [list(row) for row in workbook["infra_com"].iter_rows(values_only=True)]
    headers = [normalize(value) for value in rows[0]]
    name_idx = column_index(headers, "taxon")
    cat_idx = next(index for index, value in enumerate(headers) if value.startswith("lr"))
    for row in rows[1:]:
        category = normalize_category(row[cat_idx] if cat_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        if not category or not name:
            continue
        out.append({"code": None, "name": name, "category": category})
    return out


def parse_herpeto(path: Path, sheet_name: str) -> list[dict]:
    rows = workbook_rows(path, sheet_name)
    header_row = find_header_row(rows, ["nom scientifique", "categorie liste rouge paca"])
    headers = [normalize(value) for value in rows[header_row]]
    name_idx = column_index(headers, "nom scientifique")
    cat_idx = column_index(headers, "categorie liste rouge paca")
    out = []
    for row in rows[header_row + 1 :]:
        category = normalize_category(row[cat_idx] if cat_idx is not None and cat_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        if not category or not name:
            continue
        out.append({"code": None, "name": name, "category": category})
    return out


def parse_orthopteres(path: Path) -> list[dict]:
    rows = workbook_rows(path, "Ortho_LR-PACA_2018")
    header_row = find_header_row(rows, ["nom scientifique", "lr paca 2018"])
    headers = [normalize(value) for value in rows[header_row]]
    name_idx = column_index(headers, "nom scientifique")
    cat_idx = column_index(headers, "lr paca 2018")
    out = []
    for row in rows[header_row + 1 :]:
        category = normalize_category(row[cat_idx] if cat_idx is not None and cat_idx < len(row) else None)
        name = clean(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        if not category or not name:
            continue
        out.append({"code": None, "name": name, "category": category})
    return out


def parse_source(source: dict, path: Path) -> list[dict]:
    key = source["key"]
    if key == "oiseaux":
        return parse_oiseaux(path)
    if key == "odonates":
        return parse_odonates(path)
    if key == "papillons":
        return parse_papillons(path)
    if key == "flore":
        return parse_flore(path)
    if key in {"amphibiens", "reptiles"}:
        return parse_herpeto(path, source["sheet"])
    if key == "orthopteres":
        return parse_orthopteres(path)
    raise RuntimeError(f"Parser inconnu: {key}")


def taxref_lookup(path: Path, wanted_codes: set[int], wanted_names: set[str], realms: set[str]):
    by_cd_nom: dict[int, tuple[int, str, str, str, str, bool]] = {}
    by_cd_ref: dict[int, tuple[int, str, str, str, str, bool]] = {}
    by_name: dict[str, set[tuple[int, str, str, str, str, bool]]] = defaultdict(set)
    by_core: dict[str, set[tuple[int, str, str, str, str, bool]]] = defaultdict(set)

    def core(name: str) -> str:
        tokens = strip_authorship(name).replace("×", "x").split()
        if len(tokens) < 2:
            return normalize(name)
        keep = tokens[:2]
        if len(tokens) >= 4 and tokens[2].casefold() in {"subsp.", "subsp", "ssp.", "ssp", "var.", "var"}:
            keep.extend(tokens[2:4])
        return normalize(" ".join(keep))

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cd_nom = as_int(row.get("CD_NOM"))
            cd_ref = as_int(row.get("CD_REF"))
            if not cd_nom or not cd_ref:
                continue
            realm = {"animalia": "fauna", "plantae": "flora"}.get(normalize(row.get("REGNE")))
            if realm not in realms:
                continue
            rank = clean(row.get("RANG")).upper()
            order = clean(row.get("ORDRE"))
            classe = clean(row.get("CLASSE"))
            accepted = cd_nom == cd_ref
            entry = (cd_ref, realm, rank, order, classe, accepted)
            if cd_nom in wanted_codes:
                by_cd_nom[cd_nom] = entry
            if cd_ref in wanted_codes and accepted:
                by_cd_ref[cd_ref] = entry
            for field in ("LB_NOM", "NOM_COMPLET", "NOM_VALIDE"):
                label = clean(row.get(field))
                if not label:
                    continue
                key = normalize(label)
                core_key = core(label)
                if key in wanted_names:
                    by_name[key].add(entry)
                if core_key in wanted_names:
                    by_core[core_key].add(entry)
    return by_cd_nom, by_cd_ref, by_name, by_core


def pick_unique(matches: set[tuple], *, taxref_order: str | None = None, taxref_class: str | None = None):
    filtered = set(matches)
    if taxref_order:
        order_key = normalize(taxref_order)
        narrowed = {entry for entry in filtered if normalize(entry[3]) == order_key}
        if narrowed:
            filtered = narrowed
    if taxref_class:
        class_key = normalize(taxref_class)
        narrowed = {entry for entry in filtered if normalize(entry[4]) == class_key}
        if narrowed:
            filtered = narrowed
    if len(filtered) == 1:
        return next(iter(filtered))
    accepted_species = {entry for entry in filtered if entry[2] == "ES" and entry[5]}
    if len(accepted_species) == 1:
        return next(iter(accepted_species))
    if len({entry[0] for entry in accepted_species}) == 1 and accepted_species:
        return next(iter(accepted_species))
    species = {entry for entry in filtered if entry[2] == "ES"}
    if len(species) == 1:
        return next(iter(species))
    refs = {entry[0] for entry in filtered}
    if len(refs) == 1:
        if accepted_species:
            return next(iter(accepted_species))
        if species:
            return next(iter(species))
        return next(iter(filtered))
    return None


def resolve(
    candidate,
    by_cd_nom,
    by_cd_ref,
    by_name,
    by_core,
    expected_realm: str,
    *,
    taxref_order: str | None = None,
    taxref_class: str | None = None,
):
    code = candidate["code"]
    if code and code in by_cd_nom:
        cd_ref, realm, rank, _order, _classe, _accepted = by_cd_nom[code]
        if realm != expected_realm:
            return None, "excluded_realm"
        if rank and rank not in SEARCHABLE_RANKS:
            return None, "excluded_rank"
        return cd_ref, "cd_nom"
    if code and code in by_cd_ref:
        cd_ref, realm, rank, _order, _classe, _accepted = by_cd_ref[code]
        if realm != expected_realm:
            return None, "excluded_realm"
        if rank and rank not in SEARCHABLE_RANKS:
            return None, "excluded_rank"
        return cd_ref, "cd_ref"

    name = candidate["name"]
    if not name:
        return None, "unmatched"
    keys = [normalize(name), normalize(strip_authorship(name))]
    tokens = strip_authorship(name).replace("×", "x").split()
    if len(tokens) >= 2:
        keys.append(normalize(" ".join(tokens[:2])))
        if len(tokens) >= 4 and tokens[2].casefold() in {"subsp.", "subsp", "ssp.", "ssp", "var.", "var"}:
            keys.append(normalize(" ".join(tokens[:4])))

    for key in keys:
        matches = {
            entry
            for entry in by_name.get(key, set()) | by_core.get(key, set())
            if entry[1] == expected_realm and (not entry[2] or entry[2] in SEARCHABLE_RANKS)
        }
        picked = pick_unique(matches, taxref_order=taxref_order, taxref_class=taxref_class)
        if picked is not None:
            return picked[0], "name"
        if len(matches) > 1:
            return None, "ambiguous"
    return None, "unmatched"


def build_package(source, input_dir: Path, lookups, checked_at: str):
    path = input_dir / source["filename"]
    digest = sha256(path)
    if digest != source["sha256"]:
        raise RuntimeError(f"{source['id']}: SHA-256 inattendu {digest}")
    candidates = parse_source(source, path)
    by_cd_nom, by_cd_ref, by_name, by_core = lookups
    stats = Counter()
    unresolved = []
    statuses = []
    seen = set()
    for candidate in candidates:
        stats["rows"] += 1
        cd_ref, mode = resolve(
            candidate,
            by_cd_nom,
            by_cd_ref,
            by_name,
            by_core,
            source["realm"],
            taxref_order=source.get("taxrefOrder"),
            taxref_class=source.get("taxrefClass"),
        )
        if cd_ref is None:
            stats[mode] += 1
            if len(unresolved) < 40:
                unresolved.append(
                    {
                        "code": candidate["code"],
                        "taxon": candidate["name"],
                        "category": candidate["category"],
                        "reason": mode,
                    }
                )
            continue
        stats["matched"] += 1
        stats[mode] += 1
        value = candidate["category"]
        if len(value) > MAX_VALUE_LENGTH:
            stats["omittedLong"] += 1
            continue
        key = (cd_ref, value)
        if key in seen:
            continue
        seen.add(key)
        statuses.append(
            {
                "cdRef": cd_ref,
                "region": "PAC",
                "category": "red_list_regional",
                "label": "Liste rouge régionale",
                "value": value,
                "sourceId": source["id"],
                "scope": "regional",
            }
        )
    candidates_n = stats["matched"] + stats["unmatched"] + stats["ambiguous"]
    match_rate = round(stats["matched"] / candidates_n, 6) if candidates_n else 1.0
    covered = sorted({status["cdRef"] for status in statuses})
    return {
        "schemaVersion": 1,
        "source": {
            "id": source["id"],
            "name": source["name"],
            "producer": PRODUCER,
            "version": source["version"],
            "publicationYear": source["year"],
            "official": True,
            "checkedAt": checked_at,
            "sha256": digest,
            "landingPage": LANDING_URL,
            "sourceUrl": source["url"],
        },
        "replaces": [
            {
                "region": "PAC",
                "category": "red_list_regional",
                "realm": source["realm"],
                "cdRefs": covered,
            }
        ],
        "statuses": sorted(statuses, key=lambda status: (status["cdRef"], status["value"])),
        "diagnostics": {
            **{key: int(value) for key, value in stats.items()},
            "matchRate": match_rate,
            "unresolvedSample": unresolved,
            "values": dict(Counter(status["value"] for status in statuses)),
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--taxref", required=True)
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--checked-at", default=date.today().isoformat())
    parser.add_argument("--min-match-rate", type=float, default=0.97)
    args = parser.parse_args()

    wanted_codes: set[int] = set()
    wanted_names: set[str] = set()
    realms: set[str] = set()
    for source in SOURCES:
        path = Path(args.input_dir) / source["filename"]
        for candidate in parse_source(source, path):
            realms.add(source["realm"])
            if candidate["code"]:
                wanted_codes.add(candidate["code"])
            if candidate["name"]:
                wanted_names.add(normalize(candidate["name"]))
                wanted_names.add(normalize(strip_authorship(candidate["name"])))
                tokens = strip_authorship(candidate["name"]).replace("×", "x").split()
                if len(tokens) >= 2:
                    wanted_names.add(normalize(" ".join(tokens[:2])))

    lookups = taxref_lookup(Path(args.taxref), wanted_codes, wanted_names, realms)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    total = 0
    for source in SOURCES:
        package = build_package(source, Path(args.input_dir), lookups, args.checked_at)
        diagnostics = package["diagnostics"]
        print(json.dumps({"source": source["id"], **diagnostics}, ensure_ascii=False, indent=2))
        if diagnostics["matchRate"] < args.min_match_rate:
            raise SystemExit(
                f"{source['id']}: taux de raccord TAXREF insuffisant "
                f"{diagnostics['matchRate']:.2%} < {args.min_match_rate:.2%}"
            )
        if not package["statuses"]:
            raise SystemExit(f"{source['id']}: aucun statut produit")
        if "url" in package["source"]:
            raise SystemExit(f"{source['id']}: champ url interdit")
        output = out_dir / f"pac-lrr-{source['key']}.json"
        output.write_text(json.dumps(package, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
        total += len(package["statuses"])
        print(f"Paquet écrit: {output} — {len(package['statuses'])} statuts")
    print(f"PACA LRR: {len(SOURCES)} paquets, {total} statuts")


if __name__ == "__main__":
    main()
